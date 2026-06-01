"""
Generate the two template files shipped with the public release from the
working master/Master_Input.xlsx:

  sample/Master_Input_DEMO.xlsx   — pre-filled with 100% fictional data
  sample/Master_Input_EMPTY.xlsx  — blank template, all data cells cleared

This is reproducible — re-run any time the master schema changes:
    python3 sample/build_demo.py

Both files ship inside the public binary so anyone downloading the release
can pick either starting point without any real client data exposed.
"""

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook


HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(HERE)
SRC = os.path.join(PROJECT_ROOT, "master", "Master_Input.xlsx")
DST_DEMO = os.path.join(HERE, "Master_Input_DEMO.xlsx")
DST_EMPTY = os.path.join(HERE, "Master_Input_EMPTY.xlsx")

# Replace any client-specific references on the START HERE tab with neutral text.
START_HERE_REPLACEMENTS = {
    3: (1, "This file ships with two starter modes:\n"
           "  • Master_Input_DEMO.xlsx  — pre-filled with fictional demo data (ABC SAMPLE PRIVATE LIMITED)\n"
           "  • Master_Input_EMPTY.xlsx — blank template, fill in your client's data\n"
           "Rename whichever you want to use as 'Master_Input.xlsx' and place it next to the app."),
    15: (3, "Save THIS file with a client-specific name "
            "(e.g. Master_ClientName_FY2025.xlsx). Keep one copy per "
            "client per year for your records."),
}


# Fictional company. Every field below is invented — no real CIN, no real PAN.
DEMO = {
    "company": {
        "name":           "ABC SAMPLE PRIVATE LIMITED",
        "old_name":       "",
        "cin":            "U99999XX9999PTC999999",
        "address":        "Sample Building No. 0, Sample Street, Sample City - 999999, Sample State, India",
        "email":          "sample@example.com",
        "phone":          9999999999,
        "incorp":         datetime(2020, 4, 1),
        "business":       "sample business activity (replace with your client's business)",
        "agm_number":     "06th",
        "agm_date":       datetime(2026, 9, 30),
        "agm_day":        "Wednesday",
        "agm_words":      "30th September, 2026",
        "agm_time":       "11:00 A.M.",
        "agm_venue":      "Sample Building No. 0, Sample Street, Sample City - 999999, Sample State, India",
        "fy_end_text":    datetime(2026, 3, 31),
        "fy_label":       "2025-2026",
        "prev_fy_end":    datetime(2025, 3, 31),
        "curr_fy_end":    datetime(2026, 3, 31),
        "notice_date":    datetime(2026, 9, 5),
        "notice_place":   "Sample City",
        "prev_agm_date":  datetime(2025, 9, 30),
        "egm_date":       "",
        "roc_address":    "Sample Bhawan, 0/0, Sample Floor, Sample Area, Sample City - 999999",
    },
    # First / second signing director slot on Company tab. Names match
    # Directors rows 1+2 so the auto-fill stays consistent. All values are
    # generic role-based placeholders — never real human names.
    "signing": {
        "first_name":  "DIRECTOR ONE",
        "first_desg":  "Director",
        "first_din":   "10000001",
        "second_name": "DIRECTOR TWO",
        "second_desg": "Director",
        "second_din":  "10000002",
        "dp_name":     "DIRECTOR ONE",
        "dp_din":      "10000001",
    },
    "directors": [
        # (sno, din, name, address, appt, cessation, regularise, description)
        (1, "10000001", "Director One",   "Sample City", datetime(2020, 4, 1),  None, "NO",  ""),
        (2, "10000002", "Director Two",   "Sample City", datetime(2020, 4, 1),  None, "NO",  ""),
        (3, "10000003", "Director Three", "Sample City", datetime(2025, 11, 1), None, "YES",
         "Director Three (DIN: 10000003) was appointed as an Additional Director of the Company in the Board Meeting held on 01st November 2025 and will be regularised in the upcoming Annual General Meeting."),
        (4, "10000004", "Director Four",  "Sample City", datetime(2020, 4, 1),  None, "NO",  ""),
    ],
    "shareholders": [
        # (sno, type, category, name, folio, pan, occupation, gender, shares, nominal)
        (1, "Individual", "Promoter", "Promoter Alpha", 1, "AAAPA0001A", "Business",   "Male",   3000, 10),
        (2, "Individual", "Promoter", "Promoter Beta",  2, "AAAPB0002B", "Business",   "Female", 3000, 10),
        (3, "Individual", "Promoter", "Promoter Gamma", 3, "AAAPG0003C", "Business",   "Male",   2000, 10),
        (4, "Individual", "Promoter", "Promoter Delta", 4, "AAAPD0004D", "Service",    "Female", 2000, 10),
    ],
    "board_meetings": [
        datetime(2025, 4, 15),
        datetime(2025, 7, 18),
        datetime(2025, 10, 10),
        datetime(2026, 1, 20),
        datetime(2026, 3, 15),
    ],
    "financials": {
        # (label, current, previous)
        "Total Revenue":                                (5000, 4500),
        "Total Expenses":                               (4200, 4000),
        "Profit/Loss before Tax":                       (800,  500),
        "Current Tax":                                  (200,  130),
        "Deferred Tax":                                 (0,    0),
        "Excess/short provision relating to earlier tax": (0,    0),
        "Profit/Loss after Tax":                        (600,  370),
        "Amount Transferred to Reserves (INR)":         ("Nil", ""),
        "Dividend Amount (INR)":                        ("Nil", ""),
    },
    "auditor": {
        "Auditor Firm Name":                "Sample Auditors & Associates",
        "Auditor Designation":              "Chartered Accountants",
        "Firm Registration Number (FRN)":   "012345N",
        "Auditor Address":                  "Office No. 10, Sample Tower, Nariman Point, Mumbai - 400021",
        "Tenure End FY (Year of last AGM in tenure)": 2031,
        "Term Length (years)":              5,
        "Auditor Partner Name":             "Sample Partner",
        "Auditor Partner Membership Number": "123456",
    },
    "employees": {
        "Female Employees Count":   2,
        "Male Employees Count":     5,
        "Transgender Employees Count": 0,
        "Sexual Harassment Complaints Received": 0,
        "Sexual Harassment Complaints Disposed Off": 0,
        "Sexual Harassment Complaints Pending Beyond 90 Days": 0,
    },
    "share_capital": {
        # Numeric inputs — loader builds the description from these.
        "Authorized — Equity: No. of Shares":              100000,
        "Authorized — Equity: Nominal Value per Share (Rs.)": 10,
        "Authorized — Preference: No. of Shares":          0,
        "Authorized — Preference: Nominal Value per Share (Rs.)": 0,
        "Issued — Equity: No. of Shares":                  10000,
        "Issued — Equity: Nominal Value per Share (Rs.)":  10,
        "Issued — Preference: No. of Shares":              0,
        "Issued — Preference: Nominal Value per Share (Rs.)": 0,
        "Paid-up — Equity: No. of Shares":                 10000,
        "Paid-up — Equity: Nominal Value per Share (Rs.)": 10,
        "Paid-up — Preference: No. of Shares":             0,
        "Paid-up — Preference: Nominal Value per Share (Rs.)": 0,
    },
}


def _norm(s):
    """Match labels loosely — strip ' *', '(optional)', '— optional', whitespace, lowercase."""
    import re
    s = str(s or "").strip()
    s = re.sub(r"\s*\*+\s*$", "", s)
    s = re.sub(r"\s*[—\-]\s*optional\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(\s*optional\s*\)\s*$", "", s, flags=re.IGNORECASE)
    return s.strip().lower()


def _set_kv(ws, label_to_value, key_col=2, val_col=3):
    """Walk a key/value sheet; for any label that matches a key in
    label_to_value, write its value into column `val_col`.
    Returns the set of labels that were matched (for reporting)."""
    matched = set()
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=key_col).value
        if not label:
            continue
        normalized = _norm(label)
        for key, value in label_to_value.items():
            if _norm(key) == normalized:
                _safe_set(ws, r, val_col, value)
                matched.add(key)
                break
    return matched


def _strip_all_hyperlinks(wb):
    """Final pass — drop every cell hyperlink (mailto:, http://, etc.) so old
    targets from the source workbook never survive into the published file."""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.__class__.__name__ == "MergedCell":
                    continue
                if cell.hyperlink is not None:
                    cell.hyperlink = None


def _clear_rows(ws, start_row, end_row, cols):
    """Blank out specific columns in a row range without disturbing other data.
    Silently skips merged cells (their .value is read-only). Strips hyperlinks
    too so old mailto:/https:// targets don't survive."""
    for r in range(start_row, end_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.value = None
            if cell.hyperlink is not None:
                cell.hyperlink = None


def _sanitize_start_here(wb):
    """Replace any hardcoded client references on the 'START HERE' tab."""
    if "START HERE" not in wb.sheetnames:
        return
    ws = wb["START HERE"]
    for row, (col, new_text) in START_HERE_REPLACEMENTS.items():
        ws.cell(row=row, column=col).value = new_text


def build_demo():
    """Write sample/Master_Input_DEMO.xlsx — fully fictional 'ABC SAMPLE' data."""
    if not os.path.exists(SRC):
        raise SystemExit(f"Source master file not found: {SRC}")
    shutil.copy(SRC, DST_DEMO)
    wb = load_workbook(DST_DEMO)
    _populate_demo(wb)
    _sanitize_start_here(wb)
    _strip_all_hyperlinks(wb)
    wb.save(DST_DEMO)
    print(f"\nWrote: {DST_DEMO}")


def build_empty():
    """Write sample/Master_Input_EMPTY.xlsx — same schema, no data."""
    if not os.path.exists(SRC):
        raise SystemExit(f"Source master file not found: {SRC}")
    shutil.copy(SRC, DST_EMPTY)
    wb = load_workbook(DST_EMPTY)
    _clear_all_data(wb)
    _sanitize_start_here(wb)
    _strip_all_hyperlinks(wb)
    wb.save(DST_EMPTY)
    print(f"Wrote: {DST_EMPTY}")


def _safe_set(ws, row, col, value):
    """Write a value to a cell, ignoring merged-cell targets that are read-only.
    Also strips any hyperlink on the cell so old mailto:/https:// targets don't leak."""
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == "MergedCell":
        return False
    cell.value = value
    if cell.hyperlink is not None:
        cell.hyperlink = None
    return True


def _clear_all_data(wb):
    """Empty every user-data cell while keeping labels, headers and dropdowns intact."""
    # Company / Auditor: clear column C (values) for every label that starts with
    # a real letter (not section divider '▸').
    for tab in ("Company", "Auditor"):
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=2).value
            if not label or str(label).strip().startswith("▸"):
                continue
            _safe_set(ws, r, 3, None)

    # Tabular tabs: wipe data rows (row 5..30)
    for tab, cols in [
        ("Directors",             list(range(1, 9))),
        ("Shareholders",          list(range(1, 12))),
        ("BoardMeetings",         list(range(1, 13))),
        ("RelatedParty",          list(range(1, 15))),
        ("AuditorRemarks",        list(range(1, 5))),
        ("MaterialChanges",       list(range(1, 5))),
        ("ShareCapitalChanges",   list(range(1, 4))),
        ("EGMMeetings",           list(range(1, 5))),
        ("BusinessNatureChanges", list(range(1, 4))),
        ("DirChanges",            list(range(1, 8))),
    ]:
        if tab in wb.sheetnames:
            _clear_rows(wb[tab], 5, 30, cols)

    # Financials: clear current+previous columns (C, D)
    if "Financials" in wb.sheetnames:
        ws = wb["Financials"]
        for r in range(5, ws.max_row + 1):
            _safe_set(ws, r, 3, None)
            _safe_set(ws, r, 4, None)

    # Toggles: clear values (column C) — user fills explicitly
    if "Toggles" in wb.sheetnames:
        ws = wb["Toggles"]
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=2).value
            if k and not str(k).strip().startswith("▸") and str(k).strip() != "Toggle":
                _safe_set(ws, r, 3, None)


def _populate_demo(wb):

    # -------- Company tab --------
    if "Company" in wb.sheetnames:
        ws = wb["Company"]
        company_kv = {
            "Company Name (full legal name)": DEMO["company"]["name"],
            "Old Name (if changed)":          DEMO["company"]["old_name"],
            "CIN":                            DEMO["company"]["cin"],
            "Registered Office Address":      DEMO["company"]["address"],
            "Email ID":                       DEMO["company"]["email"],
            "Contact Number":                 DEMO["company"]["phone"],
            "Website":                        "",
            "Date of Incorporation":          DEMO["company"]["incorp"],
            "Main Business Activity":         DEMO["company"]["business"],
            "AGM Number (1st / 2nd / 3rd...)": DEMO["company"]["agm_number"],
            "AGM Date":                       DEMO["company"]["agm_date"],
            "AGM Day":                        DEMO["company"]["agm_day"],
            "AGM Date in Words":              DEMO["company"]["agm_words"],
            "AGM Time":                       DEMO["company"]["agm_time"],
            "AGM Venue":                      DEMO["company"]["agm_venue"],
            "Financial Year End Date":        DEMO["company"]["fy_end_text"],
            "Financial Year Label":           DEMO["company"]["fy_label"],
            "Previous FY End Date":           DEMO["company"]["prev_fy_end"],
            "Current FY End Date":            DEMO["company"]["curr_fy_end"],
            "Notice Dispatch Date":           DEMO["company"]["notice_date"],
            "Notice Dispatch Place":          DEMO["company"]["notice_place"],
            "Previous AGM Date":              DEMO["company"]["prev_agm_date"],
            "EGM Date":                       DEMO["company"]["egm_date"],
            "RoC Address (for Designated Person letter)": DEMO["company"]["roc_address"],
            "First Signing Director — Name":  DEMO["signing"]["first_name"],
            "First Signing Director — Designation": DEMO["signing"]["first_desg"],
            "First Signing Director — DIN":   DEMO["signing"]["first_din"],
            "Second Signing Director — Name": DEMO["signing"]["second_name"],
            "Second Signing Director — Designation": DEMO["signing"]["second_desg"],
            "Second Signing Director — DIN":  DEMO["signing"]["second_din"],
            "Designated Person Name (Rule 9)": DEMO["signing"]["dp_name"],
            "Designated Person DIN":          DEMO["signing"]["dp_din"],
            **{k: v for k, v in DEMO["share_capital"].items()},
            **{k: v for k, v in DEMO["employees"].items()},
            # Clear any typed descriptions so the loader rebuilds from numeric inputs.
            "Authorized Capital Description": None,
            "Issued Capital Description":     None,
            "Subscribed and Paid-up Capital": None,
            "Additional Capital Description": "",
        }
        matched = _set_kv(ws, company_kv)
        print(f"Company: matched {len(matched)} labels")

    # -------- Directors tab --------
    if "Directors" in wb.sheetnames:
        ws = wb["Directors"]
        # Clear rows 5..14 (the data area below the header at row 4)
        _clear_rows(ws, 5, 14, [1, 2, 3, 4, 5, 6, 7, 8])
        for i, (sno, din, name, addr, appt, cessation, regularise, desc) in enumerate(DEMO["directors"]):
            r = 5 + i
            ws.cell(row=r, column=1).value = sno
            ws.cell(row=r, column=2).value = din
            ws.cell(row=r, column=3).value = name
            ws.cell(row=r, column=4).value = addr
            ws.cell(row=r, column=5).value = appt
            ws.cell(row=r, column=6).value = cessation
            ws.cell(row=r, column=7).value = regularise
            ws.cell(row=r, column=8).value = desc
        print(f"Directors: wrote {len(DEMO['directors'])} rows")

    # -------- Shareholders tab --------
    if "Shareholders" in wb.sheetnames:
        ws = wb["Shareholders"]
        _clear_rows(ws, 5, 16, list(range(1, 12)))
        for i, (sno, typ, cat, name, folio, pan, occ, gender, shares, nominal) in enumerate(DEMO["shareholders"]):
            r = 5 + i
            ws.cell(row=r, column=1).value = sno
            ws.cell(row=r, column=2).value = typ
            ws.cell(row=r, column=3).value = cat
            ws.cell(row=r, column=4).value = name
            ws.cell(row=r, column=5).value = folio
            ws.cell(row=r, column=6).value = pan
            ws.cell(row=r, column=7).value = occ
            ws.cell(row=r, column=8).value = gender
            ws.cell(row=r, column=9).value = shares
            ws.cell(row=r, column=10).value = nominal
            ws.cell(row=r, column=11).value = shares * nominal
        print(f"Shareholders: wrote {len(DEMO['shareholders'])} rows")

    # -------- BoardMeetings tab --------
    if "BoardMeetings" in wb.sheetnames:
        ws = wb["BoardMeetings"]
        # Clear S.No + Date and any attendance markers
        _clear_rows(ws, 5, 14, list(range(1, 13)))
        for i, d in enumerate(DEMO["board_meetings"]):
            r = 5 + i
            ws.cell(row=r, column=1).value = i + 1
            ws.cell(row=r, column=2).value = d
        print(f"BoardMeetings: wrote {len(DEMO['board_meetings'])} dates")

    # -------- Financials tab --------
    if "Financials" in wb.sheetnames:
        ws = wb["Financials"]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=2).value
            if not label:
                continue
            key = _norm(label)
            for fin_label, (cur, prev) in DEMO["financials"].items():
                if _norm(fin_label) == key:
                    ws.cell(row=r, column=3).value = cur
                    ws.cell(row=r, column=4).value = prev
                    break
        print("Financials: filled")

    # -------- Auditor tab --------
    if "Auditor" in wb.sheetnames:
        ws = wb["Auditor"]
        matched = _set_kv(ws, DEMO["auditor"])
        print(f"Auditor: matched {len(matched)} labels")

    # -------- Clear gated tabular tabs (start fresh, demo simple) --------
    for tab, cols in [
        ("RelatedParty",          list(range(1, 15))),
        ("AuditorRemarks",        list(range(1, 5))),
        ("MaterialChanges",       list(range(1, 5))),
        ("ShareCapitalChanges",   list(range(1, 4))),
        ("EGMMeetings",           list(range(1, 5))),
        ("BusinessNatureChanges", list(range(1, 4))),
        ("DirChanges",            list(range(1, 8))),
    ]:
        if tab in wb.sheetnames:
            _clear_rows(wb[tab], 5, 30, cols)

    # -------- Toggles tab (sensible defaults for the demo) --------
    if "Toggles" in wb.sheetnames:
        ws = wb["Toggles"]
        demo_toggles = {
            "AuditorReappointment":           "YES",
            "RegulariseAdditionalDirector":   "YES",   # Robert Brown is being regularised
            "RPT_Required":                   "NO",
            "CompanyHasWebsite":              "NO",
            "FraudsReported":                 "NO",
            "AuditorRemarks":                 "NO",
            "Loan186Applicable":              "NO",
            "MaterialChangesPostFY":          "NO",
            "RiskMgmtPolicyExists":           "NO",
            "CSRApplicable":                  "NO",
            "HasSubsidiariesEtc":             "NO",
            "SubsidiaryStatusChange":         "NO",
            "CostRecordsApplicable":          "NO",
            "InsolvencyProceedings":          "NO",
            "OneTimeSettlement":              "NO",
            "ChangeInBusinessNature":         "NO",
            "ChangeInBoardComposition":       "YES",  # Robert appointed
            "SignificantOrdersByRegulators":  "NO",
            "MaternityActApplicable":         "YES",
            "ChangeOfNameDuringYear":         "NO",
            "ChangeInShareCapital":           "NO",
            "HeldEGM":                        "NO",
            "AuditorReportInDirReport":       "YES",
            "SecretarialAuditorRequired":     "NO",
            "CostAuditorRequired":            "NO",
            "VigilMechanismRequired":         "NO",
            "FirstDesignatedPersonDeclaration": "YES",
            "FinancialUnit":                  "Thousand",
        }
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=2).value
            if not k:
                continue
            k_str = str(k).strip()
            if k_str in demo_toggles:
                ws.cell(row=r, column=3).value = demo_toggles[k_str]
        print(f"Toggles: set {len(demo_toggles)} flags")

    # _populate_demo() ends here; caller saves.


def main():
    build_demo()
    build_empty()


if __name__ == "__main__":
    main()
