"""
build_master_input.py - Builds the user-friendly master input Excel.

Features:
  * START HERE tab with step-by-step instructions and color legend
  * Dropdowns (data validations) for ALL fixed-choice fields
  * Dropdowns for ALL YES/NO toggles - no typing
  * Color-coded sections (yellow = type, green = dropdown, grey = label)
  * Bigger row heights and clearer notes
  * Frozen header rows on data tabs (Directors, Shareholders, BoardMeetings)
  * Toggles grouped by category for clarity
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ============================================================
# Color palette - clean, professional
# ============================================================
NAVY = "1F4E78"          # main heading bar
NAVY_FONT = "FFFFFF"
SECTION = "B4C7E7"        # section divider
SECTION_FONT = "1F3864"
LABEL_BG = "F2F2F2"       # grey label
INPUT_YELLOW = "FFF9C4"   # required input
DROPDOWN_BG = "E1F5C4"    # green-ish for dropdown cells
NOTE_GREY = "808080"
START_BANNER = "548235"   # green banner for START HERE

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_HEADER = Font(name="Calibri", size=14, bold=True, color=NAVY_FONT)
F_SECTION = Font(name="Calibri", size=12, bold=True, color=SECTION_FONT)
F_LABEL = Font(name="Calibri", size=11, bold=True, color="000000")
F_INPUT = Font(name="Calibri", size=11)
F_NOTE = Font(name="Calibri", size=9, italic=True, color=NOTE_GREY)
F_BIG_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
F_INSTRUCTION = Font(name="Calibri", size=11, color="000000")


def style_header_bar(cell):
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.font = F_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_section(cell):
    cell.fill = PatternFill("solid", start_color=SECTION)
    cell.font = F_SECTION
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER


def style_label(cell):
    cell.fill = PatternFill("solid", start_color=LABEL_BG)
    cell.font = F_LABEL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = BORDER


def style_input(cell, dropdown=False):
    color = DROPDOWN_BG if dropdown else INPUT_YELLOW
    cell.fill = PatternFill("solid", start_color=color)
    cell.font = F_INPUT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = BORDER


def style_note(cell):
    cell.font = F_NOTE
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_table_header(cell):
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_table_input(cell, dropdown=False):
    color = DROPDOWN_BG if dropdown else INPUT_YELLOW
    cell.fill = PatternFill("solid", start_color=color)
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER


# ============================================================
# Reusable: drop-down list values
# ============================================================
DV_YESNO = '"YES,NO"'
DV_FINUNIT = '"Hundred,Thousand,Lakh"'
DV_DESIGNATION = '"Director,Managing Director,Whole-time Director,Independent Director"'
DV_OCCUPATION = '"Business,Service,Professional,Housewife,Student,Retired,Other"'
DV_GENDER = '"Male,Female,Transgender"'
DV_SH_TYPE = '"Individual,HUF,Body Corporate,Partnership Firm,LLP,Trust,Other"'
DV_SH_CATEGORY = '"Promoter,Public,Promoter Group,Other"'
DV_ATTENDANCE = '"NA"'
DV_RPT_NATURE = '"Sale of goods,Purchase of goods,Sale of services,Purchase of services,Sale and purchase of goods or services,Lease/Renting,Loan/Advance,Other"'
DV_RPT_RELATION = '"Director is partner,Director is proprietor,Common Director,Relative of Director,Subsidiary,Holding Company,Associate Company,Other"'
DV_CAP_TYPE = '"Equity,Preference,Both"'
DV_NAME_TITLE = '"M/s,Mr.,Mrs.,Ms.,Miss,Dr.,Shri,Smt."'


def add_dv(ws, formula, cell_range_or_list):
    """Add data validation (dropdown) to a cell range OR a list of cells."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Please pick a value from the dropdown."
    dv.errorTitle = "Invalid entry"
    if isinstance(cell_range_or_list, list):
        for cell in cell_range_or_list:
            dv.add(cell)
    else:
        dv.add(cell_range_or_list)
    ws.add_data_validation(dv)


# ============================================================
# Build the workbook
# ============================================================
wb = Workbook()


# ===================================================================
# Tab 0: START HERE (full self-contained user guide)
# ===================================================================
ws = wb.active
ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 90

F_H1 = Font(name="Calibri", size=14, bold=True, color="1F3864")
F_H2 = Font(name="Calibri", size=12, bold=True, color="1F3864")
F_TAG = Font(name="Calibri", size=11, bold=True, color="C00000")
F_NORMAL = F_INSTRUCTION
ALIGN_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top")

def _h1(row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=text).font = F_H1
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26

def _h2(row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=text).font = F_H2
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def _para(row, text, height=22):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=text).font = F_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = height

def _bullet(row, text, height=22):
    ws.cell(row=row, column=1, value="•").font = Font(name="Calibri", size=12, bold=True, color="C00000")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=text).font = F_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = height

def _check(row, text, height=22):
    ws.cell(row=row, column=1, value="✓").font = Font(name="Calibri", size=12, bold=True, color="548235")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=text).font = F_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = height

def _step(row, tag, text, height=28):
    ws.cell(row=row, column=1, value="").alignment = ALIGN_TOP_LEFT
    ws.cell(row=row, column=2, value=tag).font = F_TAG
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", indent=1)
    ws.cell(row=row, column=3, value=text).font = F_NORMAL
    ws.cell(row=row, column=3).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = height

def _tab_row(row, tab, mandatory, desc, height=30):
    ws.cell(row=row, column=1, value=mandatory).font = Font(
        name="Calibri", size=10, bold=True,
        color="C00000" if mandatory == "REQ" else "548235",
    )
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=2, value=tab).font = Font(name="Calibri", size=11, bold=True, color="1F3864")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", indent=1)
    ws.cell(row=row, column=3, value=desc).font = F_NORMAL
    ws.cell(row=row, column=3).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = height

# Big banner
ws.merge_cells("A1:C2")
ws["A1"] = "AGM FINAL SET — Master Input"
ws["A1"].fill = PatternFill("solid", start_color=START_BANNER)
ws["A1"].font = F_BIG_TITLE
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 24

# Sub-banner: pre-filled sample callout
ws.merge_cells("A3:C3")
ws["A3"] = ("This file is PRE-FILLED with fictional sample data (ABC SAMPLE PRIVATE LIMITED). "
            "Open each tab and overwrite the yellow/green cells with YOUR client's data — "
            "everything is editable. Save a separate copy per client.")
ws["A3"].fill = PatternFill("solid", start_color="FFF2CC")
ws["A3"].font = Font(name="Calibri", size=11, italic=True, color="7F6000")
ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ws["A3"].border = BORDER
ws.row_dimensions[3].height = 40

row = 5

# ---- Section 1: What this tool does ----
_h1(row, "What this tool does"); row += 1
_para(row, ("Generates the COMPLETE AGM Final Set as a single Microsoft Word document — Notice of AGM, "
            "Director's Report (28 sections), Form AOC-2, Shareholders/Directors lists, Attendance sheets, "
            "and all required Resolutions + Intimation Letter. You fill this Excel; the script produces "
            "Final_Set.docx ready to print, sign and convert to PDF for ROC."), height=58); row += 2

# ---- Section 2: How to run ----
_h1(row, "How to generate the Word file"); row += 1
_para(row, ("From the agm_automation folder, double-click one of these:"), height=20); row += 1
_bullet(row, "WINDOWS:  double-click  Generate Final Set.bat"); row += 1
_bullet(row, "MAC:  double-click  Generate Final Set.command"); row += 1
_para(row, ("A small console window opens, shows progress, and stays open until you press a key. "
            "When it finishes, the Word file is in the output folder."), height=32); row += 2

# ---- Section 3: Step-by-step ----
_h1(row, "Step-by-step workflow"); row += 1
ws.column_dimensions["B"].width = 14   # narrow STEP column
steps = [
    ("STEP 1", "Save THIS file with a client-specific name (e.g. Master_ClientName_FY2025.xlsx). Keep one copy per client per year for your records."),
    ("STEP 2", "Go through each tab from left to right. Replace the sample yellow/green values with your client's data."),
    ("STEP 3", "On the Toggles tab, set every YES/NO switch to match the client's situation. Most defaults already match a typical small private company."),
    ("STEP 4", "Fill RelatedParty rows ONLY if Toggles!RPT_Required = YES. Add one row per related-party transaction."),
    ("STEP 5", "On the Directors tab, set 'Regularise?' = YES for each Additional Director being regularized at this AGM. (Also set Toggles!RegulariseAdditionalDirector = YES.)"),
    ("STEP 6", "Fill Auditor tab ONLY if Toggles!AuditorReappointment = YES."),
    ("STEP 7", "Save the Excel file (Ctrl+S)."),
    ("STEP 8", "Double-click 'Generate Final Set.bat' (Windows) or 'Generate Final Set.command' (Mac) in the agm_automation folder."),
    ("STEP 9", "Open output/Final_Set.docx in Word — review, insert the Route Map image where the placeholder is, sign, and Save As PDF."),
]
for tag, text in steps:
    _step(row, tag, text, height=34); row += 1
row += 1

# ---- Section 4: Tab-by-tab guide ----
_h1(row, "What each tab is for"); row += 1
ws.column_dimensions["B"].width = 18
# Header row
ws.cell(row=row, column=1, value="Need").font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=1).fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=row, column=1).border = BORDER
ws.cell(row=row, column=2, value="Tab").font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=row, column=2).border = BORDER
ws.cell(row=row, column=3, value="What goes in it").font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=3).fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=row, column=3).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

tab_rows = [
    ("REQ", "Company",             "Company name, CIN, registered office, AGM date/time/venue, signing directors (2), share capital description, employee counts. Fill Date of Incorporation if this is the company's 1st AGM."),
    ("REQ", "Directors",           "List of ALL directors as on 31st March. One row per director. DIN, name, address, appointment date. Leave 'Date of Cessation' blank if still a director."),
    ("REQ", "Shareholders",        "List of ALL shareholders as on 31st March. One row each — Type, Category, Name, Folio, PAN, No. of Shares, Nominal Value, Total Amount."),
    ("REQ", "BoardMeetings",       "Dates of board meetings during the FY (one row each). Attendance columns are now blank — the printable Attendance Sheet is signed by hand. Mark NA only if a director was not yet on the board on that date."),
    ("REQ", "Financials",          "Total Revenue, Expenses, current/deferred/earlier tax, and amounts for previous year. Profit/Loss rows auto-calculate. Also: amount transferred to Reserves (or 'Nil') and Dividend amount (or 'Nil')."),
    ("REQ", "Toggles",             "YES/NO switches that turn each conditional clause on or off. Most companies will use the defaults. See the cheat sheet below."),
    ("OPT", "RelatedParty",        "Fill only if Toggles!RPT_Required = YES. Add ONE ROW per related-party transaction. Each row produces its own Notice item, Explanatory Statement, AOC-2 block and Resolution."),
    ("REQ", "Auditor",             "Auditor firm name, FRN, address, tenure-end FY (e.g. 2030 for a 5-year term ending FY 2029-30). Used by auditor reappointment + Director's Report 28(5) + Intimation Letter."),
    ("OPT", "AuditorRemarks",      "Fill only if Toggles!AuditorRemarks = YES. Add ONE ROW per auditor qualification / reservation / adverse remark. Each row appears in the Director's Report (Point 8) table."),
    ("OPT", "MaterialChanges",     "Fill only if Toggles!MaterialChangesPostFY = YES. Add ONE ROW per material change/commitment between FY-end and the date of this report. Each row appears in the Director's Report (Point 13) table."),
    ("OPT", "ShareCapitalChanges", "Fill only if Toggles!ChangeInShareCapital = YES. Add ONE ROW per change in share capital during the FY. Each row appears in the Director's Report (Section 28 → Share Capital) table."),
    ("OPT", "BusinessNatureChanges","Fill only if Toggles!ChangeInBusinessNature = YES. Add ONE ROW per change in the nature of business. Each description appears in the Director's Report (Section 20) table."),
    ("OPT", "EGMMeetings",          "Fill only if Toggles!HeldEGM = YES. Add ONE ROW per Extra-Ordinary General Meeting held during the FY. Each row appears in the Director's Report (Section 28(3)(b)) table."),
]
for need, tab, desc in tab_rows:
    _tab_row(row, tab, need, desc, height=42)
    row += 1
row += 1

# ---- Section 5: Toggles cheat sheet ----
_h1(row, "Toggles — what each switch does"); row += 1
_para(row, ("All toggles are YES/NO. Defaults match a typical small private limited company. "
            "Flip a toggle and the corresponding paragraph appears or disappears in the output."), height=32); row += 1

ws.cell(row=row, column=1, value="").fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=2, value="Toggle").font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=2).fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=row, column=2).border = BORDER
ws.cell(row=row, column=3, value="When to set YES").font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ws.cell(row=row, column=3).fill = PatternFill("solid", start_color=NAVY)
ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=row, column=3).border = BORDER
ws.row_dimensions[row].height = 22
row += 1

toggle_help = [
    ("AuditorReappointment",          "Statutory Auditor's term is ending at this AGM and you're reappointing them. Triggers Notice item, Resolution and Intimation Letter."),
    ("RegulariseAdditionalDirector",  "An Additional Director was appointed during the year and needs to be regularized at this AGM. Mark each such director Regularise? = YES on the Directors tab."),
    ("RPT_Required",                  "Company entered into related-party transactions during the year that need shareholder approval. Add the entries on RelatedParty tab."),
    ("ChangeInShareCapital",          "Share capital changed during the FY (e.g., increase in authorized capital, fresh issue). If YES, add one row per change on the ShareCapitalChanges tab. Otherwise NO."),
    ("CompanyHasWebsite",             "Company has an active website (printed in Director's Report 28). If YES, fill the Website field on Company tab."),
    ("FinancialUnit",                 "Choose Hundred / Thousand / Lakh — the unit shown next to the financial figures table."),
    ("AuditorRemarks",                "Auditor's report has qualifications, reservations or adverse remarks. If YES, add one row per remark on the AuditorRemarks tab — each row appears in Director's Report Point 8."),
    ("Loan186Applicable",             "Section 186 (loans/guarantees/investments) applies to the company. Usually NO for small private companies."),
    ("ChangeInBusinessNature",        "Company's main business activity changed during the FY. If YES, add one row per change on the BusinessNatureChanges tab."),
    ("ChangeInBoardComposition",      "A director was appointed, resigned, or otherwise ceased during the year. Independent of regularization."),
    ("MaterialChangesPostFY",         "Material events between FY-end (31 Mar) and the date of this Director's Report (e.g. major contract, asset sale). If YES, add one row per event on the MaterialChanges tab."),
    ("HasSubsidiariesEtc",            "Company has any subsidiary, associate or JV. Usually NO for small private companies."),
    ("CostRecordsApplicable",         "[Pt 18(iv)] Cost records under Section 148 are applicable to the company and being maintained. Default NO."),
    ("SignificantOrdersByRegulators", "A regulator/court/tribunal passed a significant order affecting the company during the year."),
    ("ChangeOfNameDuringYear",        "Company changed its legal name during the FY. If YES, fill 'Old Name' on Company tab."),
    ("MaternityActApplicable",        "Company has female employees and the Maternity Benefit Act, 1961 applies."),
    ("HeldEGM",                       "Company held an Extra-Ordinary General Meeting during the year."),
    ("SecretarialAuditorRequired",    "Section 204 requires a Secretarial Auditor. Usually NO for small private companies."),
    ("CostAuditorRequired",           "Section 148 requires a Cost Auditor. Usually NO."),
    ("AuditorReportInDirReport",      "Include the Statutory Auditor sub-section inside Director's Report 28(5). Usually YES."),
    ("FirstDesignatedPersonDeclaration", "Set YES the FIRST time the company files the Rule 9(7) Designated Person declaration. After that filing, set NO for all future AGMs. Independent of 1st-AGM status — Rule 9(7) was inserted in 2023, so older companies file it once when they first comply."),
]
for key, when_yes in toggle_help:
    ws.cell(row=row, column=2, value=key); style_label(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=when_yes).font = F_NORMAL
    ws.cell(row=row, column=3).alignment = ALIGN_TOP_WRAP
    ws.cell(row=row, column=3).border = BORDER
    ws.row_dimensions[row].height = 32
    row += 1
row += 1

# ---- Section 6: Multi-entry tabs ----
_h1(row, "Multi-row optional tabs (gated by Toggles)"); row += 1
_para(row, ("These tabs are TABULAR — add as many rows as you need. Each tab is read only when its toggle is YES:"), height=22); row += 1
_bullet(row, ("RelatedParty (Toggles!RPT_Required) — one row per related-party transaction. Each row produces a separate Notice item, "
              "Explanatory Statement, AOC-2 block, and Certified True Copy resolution."), height=42); row += 1
_bullet(row, ("Directors tab Regularise? = YES (Toggles!RegulariseAdditionalDirector) — set YES on each Additional Director row. "
              "Each marked director produces a separate Notice item, Explanatory Statement, and Certified True Copy resolution."), height=42); row += 1
_bullet(row, ("AuditorRemarks (Toggles!AuditorRemarks) — one row per auditor qualification / reservation / adverse remark. "
              "Each row appears in the Director's Report Point 8 table."), height=42); row += 1
_bullet(row, ("MaterialChanges (Toggles!MaterialChangesPostFY) — one row per material change/commitment between FY-end and report date. "
              "Each row appears in the Director's Report Point 13 table."), height=42); row += 1
_bullet(row, ("ShareCapitalChanges (Toggles!ChangeInShareCapital) — one row per share capital change during the FY. "
              "Each row appears in the Director's Report Section 28 → Share Capital."), height=42); row += 1
_bullet(row, ("BusinessNatureChanges (Toggles!ChangeInBusinessNature) — one row per change in the nature of business. "
              "Each row appears in the Director's Report Section 20 table."), height=42); row += 1
_bullet(row, ("EGMMeetings (Toggles!HeldEGM) — one row per Extra-Ordinary General Meeting during the FY. "
              "Each row appears in the Director's Report Section 28(3)(b) table."), height=42); row += 1
_bullet(row, ("If you need more rows than the template provides, just type into the row below the last filled row — "
              "the loader picks up everything until it hits a blank row."), height=32); row += 2

# ---- Section 7: Reusing for other clients ----
_h1(row, "Using this for multiple clients"); row += 1
_check(row, ("Save a SEPARATE copy of Master_Input.xlsx per client per year, named like "
             "'Master_<ClientShortName>_<FY>.xlsx'. Keep them in a 'masters/' archive folder."), height=42); row += 1
_check(row, ("Before running for a new client, REPLACE the file at master/Master_Input.xlsx with that client's "
             "saved copy (rename your archive copy back to Master_Input.xlsx, or copy-paste it)."), height=42); row += 1
_check(row, ("Or, advanced: edit the bottom of generate_final_set.py to point to whichever input file you want."), height=22); row += 1
_check(row, ("Date format everywhere is DD/MM/YYYY (e.g., 30/09/2025). Excel may show dates in its own locale "
             "format — type as text if Excel keeps converting."), height=32); row += 1
_check(row, ("For amounts, just type numbers (e.g., 13027.70). The script auto-formats with commas in the output."), height=22); row += 2

# ---- Section 8: Cell colour guide ----
_h1(row, "Cell colour guide"); row += 1
legend = [
    ("Yellow", INPUT_YELLOW, "Type a value here"),
    ("Green",  DROPDOWN_BG,  "Click and pick from a dropdown list — do not type"),
    ("Grey",   LABEL_BG,     "Field label — don't edit"),
    ("Blue",   NAVY,         "Section / table header — don't edit"),
]
for name, color, desc in legend:
    ws.cell(row=row, column=1, value="").fill = PatternFill("solid", start_color=color)
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value=name).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=row, column=3, value=desc).font = F_NORMAL
    ws.cell(row=row, column=3).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = 22
    row += 1
row += 1

# ---- Section 9: Troubleshooting ----
_h1(row, "Troubleshooting"); row += 1
trouble = [
    ("Output document missing a section",
     "Check the Toggles tab — the corresponding switch is set to NO. Flip it to YES and regenerate."),
    ("Auditor reappointment block not appearing",
     "Set Toggles!AuditorReappointment = YES and confirm Auditor tab is filled (firm name, FRN, tenure-end FY)."),
    ("RPT block missing although I filled RelatedParty",
     "Set Toggles!RPT_Required = YES. The toggle gates the entire RPT chain (Notice item, AOC-2, Resolution)."),
    ("Regularization missing although I marked Regularise? = YES",
     "Set Toggles!RegulariseAdditionalDirector = YES. The toggle gates the regularization items."),
    ("Item numbers in Notice look wrong",
     "Item numbers auto-renumber based on which toggles are ON. If Auditor Reappointment is OFF, special-business items start one number earlier."),
    ("Wrong company name in output",
     "Check the Company tab, save the Excel (Ctrl+S), then re-run the generator."),
    ("Document looks misaligned in Word",
     "In Word, set page size to A4 portrait. The document is tuned for A4."),
]
for q, a in trouble:
    ws.cell(row=row, column=1, value="?").font = Font(name="Calibri", size=12, bold=True, color="C00000")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=2, value=q).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row, column=2).alignment = ALIGN_TOP_WRAP
    ws.cell(row=row, column=3, value=a).font = F_NORMAL
    ws.cell(row=row, column=3).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[row].height = 50
    row += 1
row += 1

# ---- Section 10: Manual finishing ----
_h1(row, "Manual steps after generation"); row += 1
_para(row, "The script automates 99% of the document. These you do by hand in Word:", height=22); row += 1
_bullet(row, "Insert the Route Map image where the [Insert image of route map] placeholder is."); row += 1
_bullet(row, "Apply digital signatures or print and physically sign."); row += 1
_bullet(row, "File → Save As → PDF for ROC filing."); row += 1
_bullet(row, "If the auditor's report had a specific qualification, fill the auditor-remarks table inside the document."); row += 1
_bullet(row, "If you have a subsidiary/JV/associate, fill the subsidiary table in section 17."); row += 1


# ===================================================================
# Tab 1: Company
# ===================================================================
ws = wb.create_sheet("Company")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 55

ws.merge_cells("A1:D1")
ws["A1"] = "COMPANY DETAILS"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:D2")
ws["A2"] = "Fill all the yellow cells with your client's data. Pick from green dropdowns. Grey labels — don't edit."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 20


def section_row(ws, row, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=label)
    style_section(ws.cell(row=row, column=1))
    ws.row_dimensions[row].height = 22


def kv_row(ws, row, label, sample_value, note, dropdown=False):
    ws.cell(row=row, column=2, value=label); style_label(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=sample_value); style_input(ws.cell(row=row, column=3), dropdown=dropdown)
    ws.cell(row=row, column=4, value=note); style_note(ws.cell(row=row, column=4))
    ws.row_dimensions[row].height = 24


# ---- Basic Info ----
row = 4
section_row(ws, row, "▸ Basic Company Details"); row += 1
basic = [
    ("Company Name (full legal name) *", "", "Used in headers, resolutions, etc."),
    ("Old Name (if changed) — optional", "", "Format: 'Formerly Known As <Old Name>'. Leave blank if no change."),
    ("CIN *", "", ""),
    ("Registered Office Address *", "", ""),
    ("Email ID *", "", ""),
    ("Contact Number *", "", ""),
    ("Website (optional)", "", "If blank, board report says 'Company doesn't have any website'"),
    ("Date of Incorporation *", "", "Format: DD/MM/YYYY"),
    ("Main Business Activity *", "",
     "Used in 'State of Company's affairs' paragraph"),
]
for label, val, note in basic:
    kv_row(ws, row, label, val, note); row += 1

# ---- AGM Details ----
row += 1
section_row(ws, row, "▸ AGM Details"); row += 1
agm = [
    ("AGM Number (1st / 2nd / 3rd...) *", "", "Format: '01st', '02nd', '03rd'..."),
    ("AGM Date *", "", "Format: DD/MM/YYYY"),
    ("AGM Day *", "", "Day of week in CAPS, e.g., TUESDAY"),
    ("AGM Date in Words *", "", "e.g., 30th September, 2025"),
    ("AGM Time *", "", "e.g., 11:00 A.M."),
    ("AGM Venue *", "", "Usually same as registered office"),
    ("Financial Year End Date *", "", "e.g., '31st March 2025'"),
    ("Financial Year Label *", "", "e.g., '2024-25' for FY 2024-2025"),
    ("Previous FY End Date *", "", "Format: DD/MM/YYYY"),
    ("Current FY End Date *", "", "Format: DD/MM/YYYY"),
    ("Notice Dispatch Date *", "", "Format: DD/MM/YYYY"),
    ("Notice Dispatch Place *", "", ""),
    ("Previous AGM Date", "", "Fill the date of the previous AGM (DD/MM/YYYY). Leave blank ONLY for a 1st AGM — that single field decides which clause appears in the doc."),
    ("EGM Date", "", "Fill ONLY if Toggles!HeldEGM = YES or the name was changed during the year. Format: DD/MM/YYYY. Used in the EGM clause and the Change-of-Name sentence."),
    ("RoC Address (for Designated Person letter) *", "",
     "Address of the Registrar of Companies for your state"),
]
agm_date_row = None
agm_day_row = None
agm_words_row = None
for label, val, note in agm:
    if label.startswith("AGM Date *"):
        agm_date_row = row
    elif label.startswith("AGM Day"):
        agm_day_row = row
    elif label.startswith("AGM Date in Words"):
        agm_words_row = row
    kv_row(ws, row, label, val, note); row += 1

# Auto-derive AGM Day + AGM Date in Words from AGM Date (real Excel date required).
# Type to override either cell — that replaces the formula.
if agm_date_row and agm_day_row and agm_words_row:
    d = f"C{agm_date_row}"
    ws.cell(row=agm_day_row, column=3,
            value=f'=IFERROR(IF({d}="","",UPPER(TEXT({d},"dddd"))),"")')
    # Ordinal suffix: 11–13 → "th"; else 1→st, 2→nd, 3→rd, others→th.
    ordinal = (
        f'IF(AND(MOD(DAY({d}),100)>=11,MOD(DAY({d}),100)<=13),"th",'
        f'CHOOSE(MOD(DAY({d}),10)+1,"th","st","nd","rd","th","th","th","th","th","th"))'
    )
    ws.cell(row=agm_words_row, column=3,
            value=f'=IFERROR(IF({d}="","",DAY({d})&{ordinal}&" "&TEXT({d},"mmmm, yyyy")),"")')
    ws.cell(row=agm_day_row, column=4, value=(
        "Auto-derived from AGM Date. Type here to override."
    ))
    ws.cell(row=agm_words_row, column=4, value=(
        "Auto-derived from AGM Date (e.g., 30th September, 2025). Type here to override."
    ))

# ---- Signing Directors ----
row += 1
section_row(ws, row, "▸ Signing Directors"); row += 1

# Capture the rows where each Signing Director field lives so the Directors-tab
# auto-fill formulas (added later) reference the right cells even if anything
# above this section shifts. NEVER hardcode these row numbers downstream.
first_signing_name_row = row
kv_row(ws, row, "First Signing Director — Name *", "", "Director who signs Notice + Intimation Letter"); row += 1
kv_row(ws, row, "First Signing Director — Designation *", "Director", "Pick from dropdown", dropdown=True)
add_dv(ws, DV_DESIGNATION, f"C{row}"); row += 1
first_signing_din_row = row
kv_row(ws, row, "First Signing Director — DIN *", "", ""); row += 1

second_signing_name_row = row
kv_row(ws, row, "Second Signing Director — Name *", "", "Co-signs Board Report, AOC-2, Resolutions"); row += 1
kv_row(ws, row, "Second Signing Director — Designation *", "Director", "Pick from dropdown", dropdown=True)
add_dv(ws, DV_DESIGNATION, f"C{row}"); row += 1
second_signing_din_row = row
kv_row(ws, row, "Second Signing Director — DIN *", "", ""); row += 1

kv_row(ws, row, "Designated Person Name (Rule 9) *", "", "Used in Director's Report 28(13). Usually one of the directors."); row += 1
kv_row(ws, row, "Designated Person DIN *", "", ""); row += 1

# ---- Share Capital ----
# Each capital category (Authorized / Issued / Paid-up) takes numeric inputs.
# The Description cell at the bottom of each block is auto-filled by the
# generator from those numbers — leave it blank to auto-generate, or type
# your own sentence to override.
row += 1
section_row(ws, row, "▸ Share Capital (as on FY end)"); row += 1

DESC_NOTE = (
    "AUTO-GENERATED from the numbers above when you run Generate Final Set. "
    "Leave blank to auto-fill. Type your own sentence here to override (auto-fill is then skipped — clear the cell to re-enable)."
)

cap_blocks = [
    ("Authorized", "Authorized Capital Description *"),
    ("Issued",     "Issued Capital Description *"),
    ("Paid-up",    "Subscribed and Paid-up Capital *"),
]
for label_prefix, desc_label in cap_blocks:
    # Sub-section header
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=2, value=f"  {label_prefix} Capital").font = Font(
        name="Calibri", size=10, bold=True, color="1F3864"
    )
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    row += 1

    kv_row(ws, row, f"{label_prefix} — Capital Type *", "Equity",
           "Equity / Preference / Both", dropdown=True)
    add_dv(ws, DV_CAP_TYPE, f"C{row}"); row += 1

    kv_row(ws, row, f"{label_prefix} — Equity: No. of Shares", "",
           "Number only, e.g., 10000. Leave blank if Capital Type = Preference."); row += 1
    kv_row(ws, row, f"{label_prefix} — Equity: Nominal Value per Share (Rs.)", "",
           "Number only, e.g., 10 or 0.50."); row += 1
    kv_row(ws, row, f"{label_prefix} — Preference: No. of Shares", "",
           "Fill only if Capital Type = Preference or Both."); row += 1
    kv_row(ws, row, f"{label_prefix} — Preference: Nominal Value per Share (Rs.)", "",
           "Fill only if Capital Type = Preference or Both."); row += 1

    kv_row(ws, row, desc_label, "", DESC_NOTE)
    ws.row_dimensions[row].height = 48
    row += 1

# Additional free-text paragraph (unchanged)
kv_row(ws, row, "Additional Capital Description (optional)", "",
       "e.g., calls in arrears, partly-paid shares, mid-year share-capital changes. Appears as a justified paragraph below the Share Capital table in the Director's Report. Leave blank to hide.")
ws.row_dimensions[row].height = 36
row += 1

# ---- Employees ----
row += 1
section_row(ws, row, "▸ Employees"); row += 1
emp = [
    ("Female Employees Count *", "", ""),
    ("Male Employees Count *", "", ""),
    ("Transgender Employees Count", "", "Leave blank if not applicable — no longer required."),
    ("Sexual Harassment Complaints Received *", "", ""),
    ("Sexual Harassment Complaints Disposed Off *", "", ""),
    ("Sexual Harassment Complaints Pending Beyond 90 Days *", "", ""),
]
female_emp_row = None
for label, val, note in emp:
    if label.startswith("Female Employees Count"):
        female_emp_row = row
    kv_row(ws, row, label, val, note); row += 1


# ===================================================================
# Tab 2: Directors
# ===================================================================
ws = wb.create_sheet("Directors")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 60
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 40

ws.merge_cells("A1:H1")
ws["A1"] = "LIST OF DIRECTORS  (as on FY End — 31st March)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
ws["A2"] = ("Add one row per director. Leave Date of Cessation blank if director is still active. Dates: DD/MM/YYYY. "
            "Rows 1–2 auto-fill DIN/Name from the Signing Directors on the Company tab. "
            "Set 'Regularise?' = YES on any Additional Director being regularized at this AGM (requires Toggles!RegulariseAdditionalDirector = YES). "
            "Description is free-text and appears in a separate table on the LIST OF DIRECTORS page of the final doc.")
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

hdr_row = 4
headers = ["S.No", "DIN", "Name of Director", "Address", "Date of Appointment", "Date of Cessation", "Regularise?", "Description"]
for i, h in enumerate(headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 30

# Blank rows ready for fill — 10 rows accommodate boards up to 10 directors.
total_dir_rows = 10
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_dir_rows):
    for c_idx in range(1, 9):
        is_dropdown = (c_idx == 7)
        style_table_input(ws.cell(row=r_idx, column=c_idx), dropdown=is_dropdown)
    # Default the Regularise column to "NO" so the dropdown has a starting value
    ws.cell(row=r_idx, column=7, value="NO")
    ws.row_dimensions[r_idx].height = 30

last_dir_row = hdr_row + total_dir_rows
add_dv(ws, DV_YESNO, f"G{hdr_row+1}:G{last_dir_row}")

# Auto-fill rows 1 & 2 from the Signing Directors on the Company tab.
# Row numbers are captured dynamically when the Signing Directors block is
# written, so adding/removing fields anywhere above (e.g. EGM Date) never
# breaks these references.
ws.cell(row=5, column=2,
        value=f'=IF(Company!C{first_signing_din_row}="","",Company!C{first_signing_din_row})')
ws.cell(row=5, column=3,
        value=f'=IF(Company!C{first_signing_name_row}="","",Company!C{first_signing_name_row})')
ws.cell(row=6, column=2,
        value=f'=IF(Company!C{second_signing_din_row}="","",Company!C{second_signing_din_row})')
ws.cell(row=6, column=3,
        value=f'=IF(Company!C{second_signing_name_row}="","",Company!C{second_signing_name_row})')

ws.freeze_panes = "A5"


# ===================================================================
# Tab 3: Shareholders
# ===================================================================
ws = wb.create_sheet("Shareholders")
ws.sheet_view.showGridLines = False
widths = [5, 14, 14, 28, 12, 14, 16, 12, 14, 14, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:K1")
ws["A1"] = "LIST OF SHAREHOLDERS  (as on FY End — 31st March)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:K2")
ws["A2"] = "Type, Category, Occupation and Gender columns have dropdowns. Total Amount = No. of Shares × Nominal Value."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

hdr_row = 4
headers = ["S.No", "Type", "Category", "Name of Shareholder", "Folio No.",
           "PAN", "Occupation", "Gender", "No. of Shares", "Nominal Value", "Total Amount (INR)"]
for i, h in enumerate(headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 36

total_rows = 12
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_rows):
    for c_idx in range(1, 12):
        is_dropdown = c_idx in (2, 3, 7, 8)
        style_table_input(ws.cell(row=r_idx, column=c_idx), dropdown=is_dropdown)

last_data_row = hdr_row + total_rows
add_dv(ws, DV_SH_TYPE, f"B{hdr_row+1}:B{last_data_row}")
add_dv(ws, DV_SH_CATEGORY, f"C{hdr_row+1}:C{last_data_row}")
add_dv(ws, DV_OCCUPATION, f"G{hdr_row+1}:G{last_data_row}")
add_dv(ws, DV_GENDER, f"H{hdr_row+1}:H{last_data_row}")

ws.freeze_panes = "A5"


# ===================================================================
# Tab 4: Board Meetings
# ===================================================================
ws = wb.create_sheet("BoardMeetings")
ws.sheet_view.showGridLines = False
# S.No, Meeting Date, then 10 director-attendance columns (matches Directors tab capacity)
NUM_DIR_COLS = 10
widths = [5, 18] + [16] * NUM_DIR_COLS
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_col_letter = get_column_letter(2 + NUM_DIR_COLS)
ws.merge_cells(f"A1:{last_col_letter}1")
ws["A1"] = "BOARD MEETINGS HELD DURING THE FINANCIAL YEAR"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_col_letter}2")
ws["A2"] = ("Add one row per board meeting. Director-name columns auto-pull from the Directors tab so headers always match. "
            "Blank attendance cells (or cells with 'N/A' / 'NA') render as empty signing boxes in the Director's Attendance sheet — type the attendance marker (e.g. 'P') in cells where the director attended.")
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

hdr_row = 4
ws.cell(row=hdr_row, column=1, value="S.No"); style_table_header(ws.cell(row=hdr_row, column=1))
ws.cell(row=hdr_row, column=2, value="Meeting Date"); style_table_header(ws.cell(row=hdr_row, column=2))
# Header columns mirror Directors!C5..C(5+N-1) — show the director's name, or fall back
# to "Director N" while the Directors tab is still empty.
for i in range(NUM_DIR_COLS):
    col = i + 3
    dir_row = 5 + i
    formula = f'=IF(Directors!C{dir_row}="","Director {i+1}",Directors!C{dir_row})'
    style_table_header(ws.cell(row=hdr_row, column=col, value=formula))
ws.row_dimensions[hdr_row].height = 30

total_bm_rows = 10
total_cols = 2 + NUM_DIR_COLS
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_bm_rows):
    for c_idx in range(1, total_cols + 1):
        is_dropdown = c_idx >= 3
        style_table_input(ws.cell(row=r_idx, column=c_idx), dropdown=is_dropdown)
    ws.row_dimensions[r_idx].height = 26

last_bm_row = hdr_row + total_bm_rows
add_dv(ws, DV_ATTENDANCE, f"C{hdr_row+1}:{last_col_letter}{last_bm_row}")

ws.freeze_panes = "A5"


# ===================================================================
# Tab 5: Financials
# ===================================================================
ws = wb.create_sheet("Financials")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 35

ws.merge_cells("A1:E1")
ws["A1"] = "FINANCIAL FIGURES"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:E2")
ws["A2"] = "Enter figures in the unit selected on Toggles!FinancialUnit. Profit/Loss rows auto-calculate."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

hdr_row = 4
ws.cell(row=hdr_row, column=2, value="Particulars"); style_table_header(ws.cell(row=hdr_row, column=2))
ws.cell(row=hdr_row, column=3, value="Year Ended (Current)"); style_table_header(ws.cell(row=hdr_row, column=3))
ws.cell(row=hdr_row, column=4, value="Year Ended (Previous)"); style_table_header(ws.cell(row=hdr_row, column=4))
ws.cell(row=hdr_row, column=5, value="Notes"); style_table_header(ws.cell(row=hdr_row, column=5))
ws.row_dimensions[hdr_row].height = 32

fin_rows = [
    ("Total Revenue", "", "", ""),
    ("Total Expenses", "", "", ""),
    ("Profit/Loss before Tax", "=C5-C6", "=D5-D6", "Auto = Revenue − Expenses"),
    ("Current Tax", "", "", ""),
    ("Deferred Tax", "", "", ""),
    ("Excess/short provision relating to earlier tax", "", "", "Leave blank if not applicable"),
    ("Profit/Loss after Tax", "=C7-C8-C9-C10", "=D7-D8-D9-D10", "Auto = PBT − all taxes"),
]
for i, (label, cur, prev, note) in enumerate(fin_rows, start=5):
    ws.cell(row=i, column=2, value=label); style_label(ws.cell(row=i, column=2))
    cc = ws.cell(row=i, column=3, value=cur); style_input(cc)
    if isinstance(cur, (int, float)):
        cc.number_format = "#,##0.00"
    pc = ws.cell(row=i, column=4, value=prev); style_input(pc)
    if isinstance(prev, (int, float)):
        pc.number_format = "#,##0.00"
    ws.cell(row=i, column=5, value=note); style_note(ws.cell(row=i, column=5))
    ws.row_dimensions[i].height = 24

row = 5 + len(fin_rows) + 1
ws.cell(row=row, column=2, value="Amount Transferred to Reserves (INR)"); style_label(ws.cell(row=row, column=2))
ws.cell(row=row, column=3, value=""); style_input(ws.cell(row=row, column=3))
ws.cell(row=row, column=5, value="Use 'Nil' if none, else amount."); style_note(ws.cell(row=row, column=5))
ws.row_dimensions[row].height = 24

row += 1
ws.cell(row=row, column=2, value="Dividend Amount (INR)"); style_label(ws.cell(row=row, column=2))
ws.cell(row=row, column=3, value=""); style_input(ws.cell(row=row, column=3))
ws.cell(row=row, column=5, value="Use 'Nil' if none, else amount."); style_note(ws.cell(row=row, column=5))
ws.row_dimensions[row].height = 24


# ===================================================================
# Tab 6: Toggles - all dropdowns, grouped
# ===================================================================
ws = wb.create_sheet("Toggles")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 65

ws.merge_cells("A1:D1")
ws["A1"] = "TOGGLES — switches that control which clauses appear"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:D2")
ws["A2"] = "Click any GREEN cell and pick from the dropdown. Don't type. The 'Effect' column tells you what each switch does."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

hdr_row = 4
ws.cell(row=hdr_row, column=2, value="Toggle"); style_table_header(ws.cell(row=hdr_row, column=2))
ws.cell(row=hdr_row, column=3, value="Value"); style_table_header(ws.cell(row=hdr_row, column=3))
ws.cell(row=hdr_row, column=4, value="Effect"); style_table_header(ws.cell(row=hdr_row, column=4))
ws.row_dimensions[hdr_row].height = 28

toggle_groups = [
    ("▸ Display Settings", [
        ("FinancialUnit", "Thousand", "Unit shown in Director's Report financial table — 'Hundred' / 'Thousand' / 'Lakh'", "fin_unit"),
        ("CompanyHasWebsite", "NO", "If NO → 'Company doesn't have any website'. If YES → website link inserted.", "yesno"),
    ]),
    ("▸ AGM Items (Special Business)", [
        ("AuditorReappointment", "NO", "Includes auditor reappointment item, resolution & intimation letter. Set YES only when re-appointing this AGM.", "yesno"),
        ("RegulariseAdditionalDirector", "NO", "Includes Additional Director regularization item & resolution. Set YES only when regularizing director(s) this AGM.", "yesno"),
        ("RPT_Required", "NO", "Includes RPT special business item, AOC-2 details & resolution. Set YES only when there are related-party transactions.", "yesno"),
    ]),
    ("▸ Director's Report — Standard Clauses", [
        ("AuditorRemarks", "NO", "If NO → 'no observations'. YES → empty table for director's comments.", "yesno"),
        ("Loan186Applicable", "NO", "Whether Section 186 (loans/guarantees/investments) applies.", "yesno"),
        ("ChangeInBusinessNature", "NO", "If NO → 'no change'. YES → placeholder for description.", "yesno"),
        ("ChangeInBoardComposition", "NO", "If NO → 'no change'. YES → composition changed.", "yesno"),
        ("ChangeInShareCapital", "NO", "If NO → 'no change in share capital'. YES → 'change occurred' line.", "yesno"),
        ("MaterialChangesPostFY", "NO", "Material changes between FY-end and report date.", "yesno"),
        ("HasSubsidiariesEtc", "NO", "Has subsidiary/JV/associate?", "yesno"),
        ("CostRecordsApplicable", "NO", "Cost records under Section 148 applicable & maintained? (Pt 18(iv)) Default NO → 'not applicable'.", "yesno"),
        ("SignificantOrdersByRegulators", "NO", "Significant orders by regulators / courts?", "yesno"),
        ("ChangeOfNameDuringYear", "NO", "Did the company change its name during the year?", "yesno"),
    ]),
    ("▸ Employee / Compliance Clauses", [
        ("MaternityActApplicable", "NO", "If NO → 'no female employee, not applicable'.", "yesno"),
        ("HeldEGM", "NO", "Held any EGM during the year?", "yesno"),
        ("SecretarialAuditorRequired", "NO", "Section 204 applicable?", "yesno"),
        ("CostAuditorRequired", "NO", "Section 148 applicable?", "yesno"),
        ("AuditorReportInDirReport", "NO", "Include auditor sub-section in Director's Report 28(5)? Set YES if reappointing or appointing auditors at this AGM.", "yesno"),
        ("FirstDesignatedPersonDeclaration", "NO", "Is this the company's FIRST Rule 9(7) Designated Person filing? YES → 'Details of Designated Person' letter is added to the AGM pack. NO → letter skipped (already on RoC record from earlier filing).", "yesno"),
    ]),
]

row = hdr_row + 1
yesno_cells = []
finunit_cells = []
maternity_row = None
regularise_row = None
for group_label, items in toggle_groups:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=group_label)
    style_section(ws.cell(row=row, column=1))
    ws.row_dimensions[row].height = 26
    row += 1
    for key, default, effect, dv_kind in items:
        ws.cell(row=row, column=2, value=key); style_label(ws.cell(row=row, column=2))
        c = ws.cell(row=row, column=3, value=default); style_input(c, dropdown=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4, value=effect); style_note(ws.cell(row=row, column=4))
        ws.row_dimensions[row].height = 26
        if dv_kind == "yesno":
            yesno_cells.append(f"C{row}")
        else:
            finunit_cells.append(f"C{row}")
        if key == "MaternityActApplicable":
            maternity_row = row
        elif key == "RegulariseAdditionalDirector":
            regularise_row = row
        row += 1

if yesno_cells:
    add_dv(ws, DV_YESNO, yesno_cells)
if finunit_cells:
    add_dv(ws, DV_FINUNIT, finunit_cells)

# Auto-flip MaternityActApplicable to YES when Company!Female Employees Count >= 1.
# Users who want to override can type NO into the cell — that replaces the formula.
if maternity_row is not None and female_emp_row is not None:
    ws.cell(row=maternity_row, column=3,
            value=f'=IF(Company!C{female_emp_row}>=1,"YES","NO")')
    ws.cell(row=maternity_row, column=4, value=(
        "Auto-set to YES when Female Employees Count on the Company tab ≥ 1. "
        "Type NO here to override."
    ))

# Auto-flip RegulariseAdditionalDirector to YES when any Directors!G row = YES.
# Directors tab Regularise? column is G, rows 5..14.
if regularise_row is not None:
    ws.cell(row=regularise_row, column=3,
            value='=IF(COUNTIF(Directors!G5:G14,"YES")>=1,"YES","NO")')
    ws.cell(row=regularise_row, column=4, value=(
        "Auto-set to YES when any director on the Directors tab has Regularise? = YES. "
        "Type NO here to override."
    ))


# ===================================================================
# Tab 7: RelatedParty (tabular — supports multiple RPTs)
# ===================================================================
ws = wb.create_sheet("RelatedParty")
ws.sheet_view.showGridLines = False
rpt_widths = [5, 14, 24, 22, 20, 22, 14, 32, 28, 16, 14, 16, 14, 10]
for i, w in enumerate(rpt_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_col_letter = get_column_letter(len(rpt_widths))
ws.merge_cells(f"A1:{last_col_letter}1")
ws["A1"] = "RELATED PARTY TRANSACTIONS  (used only if Toggles!RPT_Required = YES — add one row per RPT)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_col_letter}2")
ws["A2"] = "Add one row per related-party transaction. Each row produces its own AOC-2 block, explanatory statement and resolution."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

rpt_headers = [
    "S.No", "PAN/CIN", "Name of Related Party", "Director/KMP Related",
    "Nature of Relationship", "Nature of Contract", "Duration",
    "Salient Terms (Value during year)", "Max Aggregate Value Approved",
    "Date of Board Approval", "Advances", "Date of Special Resolution", "MGT-14 SRN",
    "Title",
]
hdr_row = 4
for i, h in enumerate(rpt_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 38

total_rpt_rows = 6
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_rpt_rows):
    for c_idx in range(1, len(rpt_headers) + 1):
        is_dropdown = c_idx in (5, 6, 14)  # Relation, Contract, Title
        style_table_input(ws.cell(row=r_idx, column=c_idx), dropdown=is_dropdown)
    ws.row_dimensions[r_idx].height = 30

last_rpt_row = hdr_row + total_rpt_rows
add_dv(ws, DV_RPT_RELATION, f"E{hdr_row+1}:E{last_rpt_row}")
add_dv(ws, DV_RPT_NATURE, f"F{hdr_row+1}:F{last_rpt_row}")
add_dv(ws, DV_NAME_TITLE, f"N{hdr_row+1}:N{last_rpt_row}")
ws.freeze_panes = "A5"


# ===================================================================
# Tab 8: Auditor
# ===================================================================
ws = wb.create_sheet("Auditor")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 50

ws.merge_cells("A1:D1")
ws["A1"] = "AUDITOR DETAILS"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

aud = [
    ("Auditor Firm Name *", "", "Used in Notice + Resolution + Intimation + Consent Letter"),
    ("Auditor Designation *", "Chartered Accountants", "Usually 'Chartered Accountants'"),
    ("Firm Registration Number (FRN) *", "", ""),
    ("Auditor Address *", "", "Used in Intimation Letter"),
    ("Tenure End FY (Year of last AGM in tenure) *", "", "e.g., '2030' for 5-year term ending FY 2029-30"),
    ("Term Length (years) *", "5", "Usually 5 for fresh appointment"),
    ("Auditor Partner Name *", "", "Signs the Consent Letter on behalf of the firm"),
    ("Auditor Partner Membership Number *", "", "ICAI M. No. of the signing partner"),
]
row = 3
for label, val, note in aud:
    ws.cell(row=row, column=2, value=label); style_label(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=val); style_input(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value=note); style_note(ws.cell(row=row, column=4))
    ws.row_dimensions[row].height = 32
    row += 1


# DirChanges tab removed — regularizations are now driven entirely by the
# Directors-tab "Regularise?" YES/NO column. The loader still tolerates a
# DirChanges tab if one exists in an older client file.


# ===================================================================
# Tab 10: AuditorRemarks (tabular — gated by Toggles!AuditorRemarks)
# ===================================================================
ws = wb.create_sheet("AuditorRemarks")
ws.sheet_view.showGridLines = False
ar_widths = [5, 46, 46]
for i, w in enumerate(ar_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_ar_col = get_column_letter(len(ar_widths))
ws.merge_cells(f"A1:{last_ar_col}1")
ws["A1"] = "AUDITOR'S REMARKS  (used only if Toggles!AuditorRemarks = YES — add one row per remark)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_ar_col}2")
ws["A2"] = "Add one row per auditor qualification / reservation / adverse remark. Each row appears in the Director's Report Point 8 table."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ar_headers = [
    "S.No",
    "Auditor's Qualification / Reservation / Adverse Remark",
    "Directors' Comments (as per Board's Report)",
]
hdr_row = 4
for i, h in enumerate(ar_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 38

total_ar_rows = 4
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_ar_rows):
    for c_idx in range(1, len(ar_headers) + 1):
        style_table_input(ws.cell(row=r_idx, column=c_idx))
    ws.row_dimensions[r_idx].height = 40

ws.freeze_panes = "A5"


# ===================================================================
# Tab 11: MaterialChanges (tabular — gated by Toggles!MaterialChangesPostFY)
# ===================================================================
ws = wb.create_sheet("MaterialChanges")
ws.sheet_view.showGridLines = False
mc_widths = [5, 18, 74]
for i, w in enumerate(mc_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_mc_col = get_column_letter(len(mc_widths))
ws.merge_cells(f"A1:{last_mc_col}1")
ws["A1"] = "MATERIAL CHANGES & COMMITMENTS  (used only if Toggles!MaterialChangesPostFY = YES — add one row per event)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_mc_col}2")
ws["A2"] = "Add one row per material change/commitment between FY-end and the date of this Director's Report. Each row appears under Point 13."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

mc_headers = [
    "S.No", "Date of Event (DD/MM/YYYY)",
    "Description of Material Change / Commitment",
]
hdr_row = 4
for i, h in enumerate(mc_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 38

total_mc_rows = 4
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_mc_rows):
    for c_idx in range(1, len(mc_headers) + 1):
        style_table_input(ws.cell(row=r_idx, column=c_idx))
    ws.row_dimensions[r_idx].height = 36

ws.freeze_panes = "A5"


# ===================================================================
# Tab 14: EGMMeetings (tabular — gated by Toggles!HeldEGM)
# ===================================================================
ws = wb.create_sheet("EGMMeetings")
ws.sheet_view.showGridLines = False
egm_widths = [5, 18, 70]
for i, w in enumerate(egm_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_egm_col = get_column_letter(len(egm_widths))
ws.merge_cells(f"A1:{last_egm_col}1")
ws["A1"] = "EXTRA-ORDINARY GENERAL MEETINGS  (used only if Toggles!HeldEGM = YES — add one row per EGM)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_egm_col}2")
ws["A2"] = "Add one row per Extra-Ordinary General Meeting held during the FY. Each row appears in the Director's Report (Section 28(3)(b)) table."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

egm_headers = ["S.No", "Date of Meeting", "Description"]
hdr_row = 4
for i, h in enumerate(egm_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 30

total_egm_rows = 4
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_egm_rows):
    for c_idx in range(1, len(egm_headers) + 1):
        style_table_input(ws.cell(row=r_idx, column=c_idx))
    ws.row_dimensions[r_idx].height = 36
ws.freeze_panes = "A5"


# ===================================================================
# Tab 13: BusinessNatureChanges (tabular — gated by Toggles!ChangeInBusinessNature)
# ===================================================================
ws = wb.create_sheet("BusinessNatureChanges")
ws.sheet_view.showGridLines = False
bnc_widths = [5, 95]
for i, w in enumerate(bnc_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_bnc_col = get_column_letter(len(bnc_widths))
ws.merge_cells(f"A1:{last_bnc_col}1")
ws["A1"] = "CHANGE IN BUSINESS NATURE  (used only if Toggles!ChangeInBusinessNature = YES — add one row per change)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_bnc_col}2")
ws["A2"] = "Add one row per change in the nature of business during the FY. Each description appears in the Director's Report (Section 20) table."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

bnc_headers = ["S.No", "Description of Change"]
hdr_row = 4
for i, h in enumerate(bnc_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 38

total_bnc_rows = 4
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_bnc_rows):
    for c_idx in range(1, len(bnc_headers) + 1):
        style_table_input(ws.cell(row=r_idx, column=c_idx))
    ws.row_dimensions[r_idx].height = 48
ws.freeze_panes = "A5"


# ===================================================================
# Tab 12: ShareCapitalChanges (tabular — gated by Toggles!ChangeInShareCapital)
# ===================================================================
ws = wb.create_sheet("ShareCapitalChanges")
ws.sheet_view.showGridLines = False
scc_widths = [5, 95]
for i, w in enumerate(scc_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last_scc_col = get_column_letter(len(scc_widths))
ws.merge_cells(f"A1:{last_scc_col}1")
ws["A1"] = "SHARE CAPITAL CHANGES  (used only if Toggles!ChangeInShareCapital = YES — add one row per change)"
style_header_bar(ws["A1"])
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_scc_col}2")
ws["A2"] = "Add one row per share capital change during the FY. Each description appears as a paragraph below the Share Capital table in the Director's Report (Section 28)."
ws["A2"].font = F_NOTE
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

scc_headers = ["S.No", "Description"]
hdr_row = 4
for i, h in enumerate(scc_headers, start=1):
    style_table_header(ws.cell(row=hdr_row, column=i, value=h))
ws.row_dimensions[hdr_row].height = 38

total_scc_rows = 4
for r_idx in range(hdr_row + 1, hdr_row + 1 + total_scc_rows):
    for c_idx in range(1, len(scc_headers) + 1):
        style_table_input(ws.cell(row=r_idx, column=c_idx))
    ws.row_dimensions[r_idx].height = 48
ws.freeze_panes = "A5"


# ===================================================================
# Save
# ===================================================================
import os
HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(HERE)
out_path = os.path.join(PROJECT_ROOT, "master", "Master_Input.xlsx")
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
