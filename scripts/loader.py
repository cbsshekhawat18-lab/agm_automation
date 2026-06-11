"""
loader.py - Reads Master_Input.xlsx into a single Python dict for the generator.

Robust to:
  * Labels with trailing decorations: '*', '(optional)', em-dashes etc.
  * Section divider rows in the Company / Toggles tabs
  * Header rows in Directors / Shareholders / BoardMeetings being at any row
"""

from openpyxl import load_workbook
from datetime import datetime, date
import re

from share_capital_text import build_description


# ============================================================
# Helpers
# ============================================================
def _norm_date(v):
    """Return DD/MM/YYYY string from whatever Excel gave us."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


def _norm_label(s):
    """Strip decorations from a label so '(full legal name) *' matches 'full legal name'."""
    if s is None:
        return ""
    s = str(s).strip()
    # Drop trailing ' *' marks
    s = re.sub(r"\s*\*+\s*$", "", s)
    # Drop ' — optional' / ' - optional' suffixes
    s = re.sub(r"\s*[—\-]\s*optional\s*$", "", s, flags=re.IGNORECASE)
    # Drop ' (optional)' suffix
    s = re.sub(r"\s*\(\s*optional\s*\)\s*$", "", s, flags=re.IGNORECASE)
    return s.strip().lower()


def _kv_pairs(ws, key_col=2, val_col=3, start_row=1, end_row=None):
    """Walk down a sheet collecting (normalized_label -> value).
    Skips section divider rows that start with '▸'."""
    out = {}
    end_row = end_row or ws.max_row
    blank_streak = 0
    for r in range(start_row, end_row + 1):
        k = ws.cell(row=r, column=key_col).value
        v = ws.cell(row=r, column=val_col).value
        if k is None and v is None:
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue
        blank_streak = 0
        if k is None:
            continue
        k_str = str(k).strip()
        if k_str.startswith("▸"):
            continue
        norm = _norm_label(k_str)
        if not norm:
            continue
        if v is None:
            out[norm] = ""
        elif isinstance(v, (int, float, datetime, date)):
            out[norm] = v
        else:
            out[norm] = str(v).strip()
    return out


def _find_table_header_row(ws, expected_first_header, max_search=10):
    """Scan first `max_search` rows for a row whose column A equals expected_first_header."""
    target = _norm_label(expected_first_header)
    for r in range(1, max_search + 1):
        v = ws.cell(row=r, column=1).value
        if v and _norm_label(str(v)) == target:
            return r
    return None


def _safe_num(v, default=0):
    """Convert to number, return default if blank/non-numeric."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _ordinal(n):
    n = int(n)
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _parse_ddmmyyyy(s):
    """Parse a DD/MM/YYYY string (the format _norm_date produces) → datetime, or None."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%d/%m/%Y")
    except (ValueError, TypeError):
        return None


def _date_in_words(s):
    """'15/03/2024' → '15th day of March, 2024'. '' if unparseable."""
    d = _parse_ddmmyyyy(s)
    if d is None:
        return ""
    return f"{_ordinal(d.day)} day of {d.strftime('%B')}, {d.year}"


def _date_dotted(s):
    """'15/03/2024' → '15.03.2024'. '' if unparseable."""
    d = _parse_ddmmyyyy(s)
    if d is None:
        return ""
    return d.strftime("%d.%m.%Y")


def _day_of_week(s):
    """'13/05/2026' → 'WEDNESDAY'. '' if unparseable."""
    d = _parse_ddmmyyyy(s)
    if d is None:
        return ""
    return d.strftime("%A").upper()


def _agm_words(s):
    """'13/05/2026' → '13th May, 2026' (matches the format the Excel formula produces)."""
    d = _parse_ddmmyyyy(s)
    if d is None:
        return ""
    return f"{_ordinal(d.day)} {d.strftime('%B')}, {d.year}"


def _date_to_fy_end_text(v):
    """If v is a datetime/date (user typed a real date), format as '31st March 2025'.
    Otherwise return v unchanged as a stripped string."""
    if isinstance(v, (datetime, date)):
        return f"{_ordinal(v.day)} {v.strftime('%B')} {v.year}"
    if v is None:
        return ""
    return str(v).strip()


def _clean_text(v, year_only=False):
    """Coerce a cell value to a clean display string.
    - None / '' → ''
    - datetime / date → 'DD/MM/YYYY' (or just YYYY when year_only=True)
    - float that is a whole number → int string (e.g. 1.0 → '1', 9314508970.0 → '9314508970')
    - everything else → str(v).strip()
    """
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return str(v.year) if year_only else v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _is_yes(v):
    return str(v or "").strip().upper() in ("YES", "Y", "TRUE", "1")


def _sno(v):
    """Display S.No cleanly — '1' not '1.0' when Excel hands back a float."""
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


# ============================================================
# Writeback: fill auto-generated share-capital descriptions back into Excel
# ============================================================
def _writeback_company_cells(path, updates):
    """Open the workbook (formulas preserved) and write the given values into
    Company-tab cells whose label (col B) matches. Skips any cell that already
    has a user value, and silently tolerates a file locked open in Excel."""
    if not any(updates.values()):
        return
    try:
        wb = load_workbook(path)
    except Exception:
        return
    ws = wb["Company"]
    label_to_row = {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b:
            label_to_row[_norm_label(str(b))] = r
    changed = False
    for label, value in updates.items():
        if not value:
            continue
        r = label_to_row.get(_norm_label(label))
        if r is None:
            continue
        if ws.cell(row=r, column=3).value:
            continue
        ws.cell(row=r, column=3).value = value
        changed = True
    if not changed:
        return
    try:
        wb.save(path)
    except PermissionError:
        print(
            f"WARNING: could not save auto-generated capital descriptions to {path}. "
            "Close the file in Excel and re-run to write them back."
        )


# ============================================================
# Main loader
# ============================================================
def load_master(path):
    wb = load_workbook(path, data_only=True)

    # ---------- Company ----------
    ws = wb["Company"]
    company_kv = _kv_pairs(ws, key_col=2, val_col=3, start_row=4)

    def cv(*keys):
        for k in keys:
            val = company_kv.get(_norm_label(k), "")
            if val != "" and val is not None:
                return val
        return ""

    company = {
        "name": cv("Company Name (full legal name)"),
        "old_name": cv("Old Name (if changed)", "Old Name (if changed - else blank)"),
        "cin": _clean_text(cv("CIN")),
        "address": cv("Registered Office Address"),
        "email": cv("Email ID"),
        "phone": _clean_text(cv("Contact Number")),
        "website": cv("Website"),
        "incorporation_date": _norm_date(cv("Date of Incorporation")),
        "business_activity": cv("Main Business Activity"),
        "roc_address": cv("RoC Address (for Designated Person letter)", "RoC Address"),
    }

    # AGM Day and Date-in-Words fall back to deriving from AGM Date if the user
    # blanked the auto-fill formula in their Excel.
    agm_date_str = _norm_date(cv("AGM Date"))
    typed_day = str(cv("AGM Day") or "").strip()
    typed_words = str(cv("AGM Date in Words") or "").strip()
    agm_day_str = typed_day or _day_of_week(agm_date_str)
    agm_words_str = typed_words or _agm_words(agm_date_str)
    agm = {
        "number": _clean_text(cv("AGM Number (1st / 2nd / 3rd...)", "AGM Number (1st / 2nd / 3rd / etc.)")),
        "date": agm_date_str,
        "day": agm_day_str,
        "date_words": agm_words_str,
        "time": str(cv("AGM Time") or "").strip(),
        "venue": str(cv("AGM Venue") or "").strip(),
        "fy_end_date": _date_to_fy_end_text(cv("Financial Year End Date")),
        "fy_label": str(cv("Financial Year Label") or "").strip(),
        "prev_fy_end": _norm_date(cv("Previous FY End Date")),
        "curr_fy_end": _norm_date(cv("Current FY End Date")),
        "notice_date": _norm_date(cv("Notice Dispatch Date")),
        "notice_place": cv("Notice Dispatch Place"),
        "prev_agm_date": _norm_date(cv("Previous AGM Date")),
        "egm_date": _norm_date(cv("EGM Date")),
    }
    # If the user blanked the formulas, write the derived strings back into the
    # cells so they see them on next Excel open.
    _writeback_company_cells(path, {
        "AGM Day": agm_day_str if not typed_day else None,
        "AGM Date in Words": agm_words_str if not typed_words else None,
    })

    signing = {
        "first": {
            "name": cv("First Signing Director — Name", "First Signing Director - Name"),
            "designation": cv("First Signing Director — Designation", "First Signing Director - Designation"),
            "din": str(cv("First Signing Director — DIN", "First Signing Director - DIN")),
        },
        "second": {
            "name": cv("Second Signing Director — Name", "Second Signing Director - Name"),
            "designation": cv("Second Signing Director — Designation", "Second Signing Director - Designation"),
            "din": str(cv("Second Signing Director — DIN", "Second Signing Director - DIN")),
        },
        "designated_person": {
            "name": cv("Designated Person Name (Rule 9)"),
            "din": str(cv("Designated Person DIN")),
        },
    }

    # ---- Share Capital ----
    # Each category has either a typed Description (user override) OR a set of
    # numeric inputs we synthesize the description from. Synthesized values are
    # written back into the Excel Description cell so the user sees them too.
    def _cap_num(prefix, key_suffix):
        return _safe_num(cv(f"{prefix} — {key_suffix}"))

    def _cap_block(prefix, desc_key):
        typed = cv(desc_key)
        if typed:
            return str(typed), False
        desc = build_description(
            _cap_num(prefix, "Equity: No. of Shares"),
            _cap_num(prefix, "Equity: Nominal Value per Share (Rs.)"),
            _cap_num(prefix, "Preference: No. of Shares"),
            _cap_num(prefix, "Preference: Nominal Value per Share (Rs.)"),
        )
        return desc, bool(desc)

    auth_desc, auth_gen = _cap_block("Authorized", "Authorized Capital Description")
    iss_desc,  iss_gen  = _cap_block("Issued",     "Issued Capital Description")
    pp_desc,   pp_gen   = _cap_block("Paid-up",    "Subscribed and Paid-up Capital")

    share_capital = {
        "authorized": auth_desc,
        "issued":     iss_desc,
        "paid_up":    pp_desc,
        "additional": cv("Additional Capital Description"),
    }

    _writeback_company_cells(path, {
        "Authorized Capital Description":  auth_desc if auth_gen else None,
        "Issued Capital Description":      iss_desc  if iss_gen  else None,
        "Subscribed and Paid-up Capital":  pp_desc   if pp_gen   else None,
    })

    employees = {
        "female": int(_safe_num(cv("Female Employees Count"))),
        "male": int(_safe_num(cv("Male Employees Count"))),
        "transgender": int(_safe_num(cv("Transgender Employees Count"))),
        "sh_received": int(_safe_num(cv("Sexual Harassment Complaints Received"))),
        "sh_disposed": int(_safe_num(cv("Sexual Harassment Complaints Disposed Off"))),
        "sh_pending": int(_safe_num(cv("Sexual Harassment Complaints Pending Beyond 90 Days"))),
    }

    # ---------- Directors ----------
    ws = wb["Directors"]
    hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
    directors = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        din = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if not (name and str(name).strip()) and not (din and str(din).strip()):
            continue
        directors.append({
            "sno": _sno(ws.cell(row=r, column=1).value),
            "din": str(din) if din is not None else "",
            "name": str(name).strip() if name else "",
            "address": str(ws.cell(row=r, column=4).value or ""),
            "appointment": _norm_date(ws.cell(row=r, column=5).value),
            "cessation": _norm_date(ws.cell(row=r, column=6).value),
            "regularise": _is_yes(ws.cell(row=r, column=7).value),
            "description": str(ws.cell(row=r, column=8).value or "").strip(),
        })

    # ---------- Shareholders ----------
    ws = wb["Shareholders"]
    hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
    shareholders = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=4).value
        if not (name and str(name).strip()):
            continue
        shareholders.append({
            "sno": _sno(ws.cell(row=r, column=1).value),
            "type": str(ws.cell(row=r, column=2).value or ""),
            "category": str(ws.cell(row=r, column=3).value or ""),
            "name": str(name).strip(),
            "folio": str(ws.cell(row=r, column=5).value or ""),
            "pan": str(ws.cell(row=r, column=6).value or ""),
            "occupation": str(ws.cell(row=r, column=7).value or ""),
            "gender": str(ws.cell(row=r, column=8).value or ""),
            "shares": _safe_num(ws.cell(row=r, column=9).value),
            "nominal": _safe_num(ws.cell(row=r, column=10).value),
            "total": _safe_num(ws.cell(row=r, column=11).value),
        })

    # ---------- BoardMeetings ----------
    # The BoardMeetings header now contains Excel formulas that pull names from
    # the Directors tab, so we don't trust the cached header value. The
    # Directors tab is the source of truth for both the names and the column
    # count (capped at the 6 attendance columns BoardMeetings provides).
    ws = wb["BoardMeetings"]
    hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
    director_short_names = [d["name"] for d in directors[:10] if d["name"]]

    board_meetings = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        d = ws.cell(row=r, column=2).value
        if not d or not str(d).strip():
            continue
        attendance = []
        for c in range(3, 3 + len(director_short_names)):
            attendance.append(str(ws.cell(row=r, column=c).value or ""))
        board_meetings.append({
            "sno": _sno(ws.cell(row=r, column=1).value),
            "date": _norm_date(d),
            "attendance": attendance,
        })

    # ---------- Financials ----------
    ws = wb["Financials"]
    fin_labels = [
        ("total_revenue", "Total Revenue"),
        ("total_expenses", "Total Expenses"),
        ("pbt", "Profit/Loss before Tax"),
        ("current_tax", "Current Tax"),
        ("deferred_tax", "Deferred Tax"),
        ("earlier_tax", "Excess/short provision relating to earlier tax"),
        ("pat", "Profit/Loss after Tax"),
    ]
    fin_lookup = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if label:
            fin_lookup[_norm_label(label)] = (
                ws.cell(row=r, column=3).value,
                ws.cell(row=r, column=4).value,
            )
    fin = {}
    for key, label in fin_labels:
        cur, prev = fin_lookup.get(_norm_label(label), (0, 0))
        fin[key] = {"current": cur if cur is not None else 0,
                    "previous": prev if prev is not None else 0}
    fin["reserves"] = str(fin_lookup.get(_norm_label("Amount Transferred to Reserves (INR)"), ("Nil", ""))[0] or "Nil")
    fin["dividend"] = str(fin_lookup.get(_norm_label("Dividend Amount (INR)"), ("Nil", ""))[0] or "Nil")

    # If the user filled inputs in Excel but Excel hasn't cached the formula
    # results (common with Sheets / LibreOffice / when the file is saved
    # programmatically), PBT and PAT come through as 0. Compute them in Python
    # as a fallback so the Director's Report doesn't show "Rs. -/-".
    def _as_num(v):
        try: return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError): return 0.0

    for col in ("current", "previous"):
        rev = _as_num(fin["total_revenue"][col])
        exp = _as_num(fin["total_expenses"][col])
        cur_tax = _as_num(fin["current_tax"][col])
        def_tax = _as_num(fin["deferred_tax"][col])
        earl_tax = _as_num(fin["earlier_tax"][col])
        # Only compute when the corresponding inputs exist; otherwise leave 0.
        if (rev or exp) and not _as_num(fin["pbt"][col]):
            fin["pbt"][col] = rev - exp
        if (rev or exp) and not _as_num(fin["pat"][col]):
            fin["pat"][col] = _as_num(fin["pbt"][col]) - cur_tax - def_tax - earl_tax

    # ---------- Toggles ----------
    # Yes/No flags are normalized to upper-case. Free-text toggles (e.g.
    # FinancialUnit = Thousand/Lakh/Hundred) are normalized to title-case so
    # they read correctly inside body text like "(In Hundred)".
    ws = wb["Toggles"]
    toggles = {}
    for r in range(5, ws.max_row + 1):
        k = ws.cell(row=r, column=2).value
        v = ws.cell(row=r, column=3).value
        if k:
            k_str = str(k).strip()
            if k_str.startswith("▸"):
                continue
            if v is None or v == "":
                toggles[k_str] = "NO"
                continue
            raw = str(v).strip()
            if raw.upper() in ("YES", "NO", "Y", "N", "TRUE", "FALSE", "1", "0"):
                toggles[k_str] = raw.upper()
            else:
                toggles[k_str] = raw.title()

    # ---------- Related Party (tabular: one row per RPT) ----------
    ws = wb["RelatedParty"]
    hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
    rpts = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        if not (name and str(name).strip()):
            continue
        rpts.append({
            "sno": _sno(ws.cell(row=r, column=1).value),
            "pan_cin": str(ws.cell(row=r, column=2).value or ""),
            "title": str(ws.cell(row=r, column=14).value or "").strip(),
            # Keep the name exactly as typed (preserves "Mr. AK Sen",
            # "M/S ABC PRIVATE LIMITED", etc.). The Title column is OPTIONAL
            # — if user typed the honorific into Name, don't double up.
            "name": str(name).strip(),
            "related_director": str(ws.cell(row=r, column=4).value or ""),
            "relationship": str(ws.cell(row=r, column=5).value or ""),
            "nature_of_contract": str(ws.cell(row=r, column=6).value or ""),
            "duration": str(ws.cell(row=r, column=7).value or ""),
            "salient_terms": str(ws.cell(row=r, column=8).value or ""),
            "max_value": str(ws.cell(row=r, column=9).value or ""),
            "board_approval_date": _norm_date(ws.cell(row=r, column=10).value),
            "advances": str(ws.cell(row=r, column=11).value or ""),
            "special_resolution_date": _norm_date(ws.cell(row=r, column=12).value),
            "mgt14_srn": str(ws.cell(row=r, column=13).value or ""),
        })

    # ---------- Auditor ----------
    ws = wb["Auditor"]
    aud_kv = _kv_pairs(ws, key_col=2, val_col=3, start_row=3)

    def av(*keys):
        for k in keys:
            val = aud_kv.get(_norm_label(k), "")
            if val != "" and val is not None:
                return val
        return ""

    auditor = {
        # Keep the firm name exactly as typed (M/s, Mr., etc. preserved).
        "firm_name": str(av("Auditor Firm Name") or "").strip(),
        "designation": av("Auditor Designation"),
        "frn": _clean_text(av("Firm Registration Number (FRN)")),
        "address": av("Auditor Address"),
        "tenure_end_fy": _clean_text(av("Tenure End FY (Year of last AGM in tenure)"), year_only=True),
        "term_years": _clean_text(av("Term Length (years)")) or "5",
        "partner_name": av("Auditor Partner Name"),
        "partner_mno": _clean_text(av("Auditor Partner Membership Number")),
    }

    # ---------- Regularizations ----------
    # Primary source: any row on the Directors tab with Regularise? = YES.
    # Secondary source (legacy / explicit overrides): rows on the DirChanges tab
    # not already covered by DIN. appt_words / appt_dotted are auto-derived from
    # the appointment date if not provided.
    regularizations = []
    seen_dins = set()

    for d in directors:
        if not d.get("regularise"):
            continue
        din = d["din"]
        regularizations.append({
            "sno": len(regularizations) + 1,
            "name": d["name"],
            "din": din,
            "appt_date": d["appointment"],
            "appt_words": _date_in_words(d["appointment"]),
            "appt_dotted": _date_dotted(d["appointment"]),
        })
        if din:
            seen_dins.add(din)

    # Legacy DirChanges tab — older client files may still have it. New
    # templates don't generate this tab; fall through silently if it's missing.
    # Belt-and-braces: if any director on the Directors tab is marked
    # Regularise? = YES, force the RegulariseAdditionalDirector toggle to YES
    # so the doc actually renders the items. The Toggles-tab formula does this
    # automatically in new templates, but older client files may have a
    # hardcoded NO; this guarantees consistency.
    if any(d.get("regularise") for d in directors):
        toggles["RegulariseAdditionalDirector"] = "YES"

    if "DirChanges" in wb.sheetnames:
        ws = wb["DirChanges"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if not (name and str(name).strip()):
                continue
            din = str(ws.cell(row=r, column=3).value or "")
            if din and din in seen_dins:
                continue
            appt_date_str = _norm_date(ws.cell(row=r, column=4).value)
            typed_words = str(ws.cell(row=r, column=5).value or "").strip()
            typed_dotted = str(ws.cell(row=r, column=6).value or "").strip()
            regularizations.append({
                "sno": len(regularizations) + 1,
                "name": str(name).strip(),
                "din": din,
                "appt_date": appt_date_str,
                "appt_words": typed_words or _date_in_words(appt_date_str),
                "appt_dotted": typed_dotted or _date_dotted(appt_date_str),
            })

    # ---------- AuditorRemarks (tabular: gated by Toggles!AuditorRemarks) ----------
    auditor_remarks = []
    if "AuditorRemarks" in wb.sheetnames:
        ws = wb["AuditorRemarks"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            remark = ws.cell(row=r, column=2).value
            if not (remark and str(remark).strip()):
                continue
            auditor_remarks.append({
                "sno": _sno(ws.cell(row=r, column=1).value),
                "remark": str(remark).strip(),
                "directors_comment": str(ws.cell(row=r, column=3).value or ""),
            })

    # ---------- MaterialChanges (tabular: gated by Toggles!MaterialChangesPostFY) ----------
    material_changes = []
    if "MaterialChanges" in wb.sheetnames:
        ws = wb["MaterialChanges"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            description = ws.cell(row=r, column=3).value
            if not (description and str(description).strip()):
                continue
            material_changes.append({
                "sno": _sno(ws.cell(row=r, column=1).value),
                "event_date": _norm_date(ws.cell(row=r, column=2).value),
                "description": str(description).strip(),
            })

    # ---------- ShareCapitalChanges (tabular: gated by Toggles!ChangeInShareCapital) ----------
    share_capital_changes = []
    if "ShareCapitalChanges" in wb.sheetnames:
        ws = wb["ShareCapitalChanges"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            description = ws.cell(row=r, column=2).value
            if not (description and str(description).strip()):
                continue
            share_capital_changes.append({
                "sno": _sno(ws.cell(row=r, column=1).value),
                "description": str(description).strip(),
            })

    # ---------- EGMMeetings (tabular: gated by Toggles!HeldEGM) ----------
    egm_meetings = []
    if "EGMMeetings" in wb.sheetnames:
        ws = wb["EGMMeetings"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            d = ws.cell(row=r, column=2).value
            if not d:
                continue
            date_str = _norm_date(d)
            if not date_str:
                continue
            egm_meetings.append({
                "sno": _sno(ws.cell(row=r, column=1).value),
                "date": date_str,
                "description": str(ws.cell(row=r, column=3).value or "").strip(),
            })

    # ---- Fallback: derive EGM rows from MaterialChanges when there's no
    # EGMMeetings tab (older master templates kept EGM info on MaterialChanges).
    # We pull any MaterialChanges row whose description mentions an EGM.
    # The doc's EGM table then renders automatically — no manual data move.
    if not egm_meetings and material_changes:
        for mc in material_changes:
            desc = mc.get("description", "") or ""
            if "extra ordinary general meeting" in desc.lower() or "extraordinary general meeting" in desc.lower() or " egm " in desc.lower():
                egm_meetings.append({
                    "sno": len(egm_meetings) + 1,
                    "date": mc.get("event_date", ""),
                    "description": desc,
                })

    # ---------- BusinessNatureChanges (tabular: gated by Toggles!ChangeInBusinessNature) ----------
    business_nature_changes = []
    if "BusinessNatureChanges" in wb.sheetnames:
        ws = wb["BusinessNatureChanges"]
        hdr_row = _find_table_header_row(ws, "S.No") or _find_table_header_row(ws, "S. No.") or 4
        for r in range(hdr_row + 1, ws.max_row + 1):
            description = ws.cell(row=r, column=2).value
            if not (description and str(description).strip()):
                continue
            business_nature_changes.append({
                "sno": _sno(ws.cell(row=r, column=1).value),
                "description": str(description).strip(),
            })

    # ---------- Auto-flip gated toggles based on filled tabular data ----------
    # If the user filled rows on a gated tab but forgot to set the YES/NO
    # toggle, force the toggle to YES so the doc actually renders that section.
    # Mirrors the Excel formulas in newer templates; keeps older client files
    # consistent. The user can still explicitly set the toggle to NO and clear
    # the rows to suppress.
    _gates = [
        ("RPT_Required",            rpts),
        ("AuditorRemarks",          auditor_remarks),
        ("MaterialChangesPostFY",   material_changes),
        ("HeldEGM",                 egm_meetings),
        ("ChangeInBusinessNature",  business_nature_changes),
        ("ChangeInShareCapital",    share_capital_changes),
    ]
    for toggle_key, rows in _gates:
        if rows:
            toggles[toggle_key] = "YES"

    return {
        "company": company,
        "agm": agm,
        "signing": signing,
        "share_capital": share_capital,
        "employees": employees,
        "directors": directors,
        "shareholders": shareholders,
        "board_meetings": board_meetings,
        "board_meeting_director_names": director_short_names,
        "financials": fin,
        "toggles": toggles,
        "rpts": rpts,
        "auditor": auditor,
        "regularizations": regularizations,
        "auditor_remarks": auditor_remarks,
        "material_changes": material_changes,
        "share_capital_changes": share_capital_changes,
        "business_nature_changes": business_nature_changes,
        "egm_meetings": egm_meetings,
    }


def fmt_num(n, decimals=2, signed=False):
    """Format number with commas. Returns '-' if zero/None.
    Default behavior: losses shown as positive (the surrounding sentence
    conveys 'Loss'). Pass signed=True for tables where the minus sign
    should be preserved (e.g. Financial Summary table)."""
    if n in (None, "", 0, 0.0):
        return "-"
    try:
        v = float(n)
        if signed:
            return f"{v:,.{decimals}f}"
        return f"{abs(v):,.{decimals}f}"
    except (ValueError, TypeError):
        return str(n)


def yes(toggles, key):
    """True if toggle is 'YES'."""
    return toggles.get(key, "NO").upper() == "YES"


if __name__ == "__main__":
    import os
    HERE = os.path.dirname(os.path.abspath(__file__))
    PROJECT_ROOT = os.path.dirname(HERE)
    data = load_master(os.path.join(PROJECT_ROOT, "master", "Master_Input.xlsx"))
    print("Company:", data["company"]["name"])
    print("CIN:", data["company"]["cin"])
    print(f"Directors loaded: {len(data['directors'])}")
    print(f"Shareholders loaded: {len(data['shareholders'])}")
    print(f"Board meetings loaded: {len(data['board_meetings'])}")
    print(f"Toggles loaded: {len(data['toggles'])}")
    print(f"RPT entries loaded: {len(data['rpts'])}")
    print(f"Regularizations loaded: {len(data['regularizations'])}")
    print("Sample financial: PAT current =", data["financials"]["pat"]["current"])
    print("Sample toggle: RPT_Required =", data["toggles"].get("RPT_Required"))
    print("Auditor:", data["auditor"]["firm_name"])
    print("First director:", data["directors"][0]["name"] if data["directors"] else "(none)")
    print("First shareholder:", data["shareholders"][0]["name"] if data["shareholders"] else "(none)")
    print("First signing:", data["signing"]["first"]["name"])
    if data["rpts"]:
        print("First RPT name:", data["rpts"][0]["name"])
    if data["regularizations"]:
        print("First regularization:", data["regularizations"][0]["name"])
