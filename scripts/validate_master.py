"""Pre-flight check for Master_Input.xlsx.

Scans the workbook for blank required fields and prints a clear list. A field
is "required" when its label on the Company / Auditor tab ends with '*', or
when a tabular tab (Directors, Shareholders, BoardMeetings) has no usable rows.

Call validate(path) before generating the doc. It hard-stops the run with
sys.exit(1) if anything is missing.
"""

import sys
from openpyxl import load_workbook


def _is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def _collect_formula_cells(wb):
    """Return a set of (sheet_name, row, col) tuples for every formula cell."""
    cells = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cells.add((sheet, cell.row, cell.column))
    return cells


# Labels that should never be flagged as missing:
#   * loader auto-derives them if blank (AGM Day, AGM Date in Words, Share Capital descriptions)
#   * or they're explicitly optional per user request (Transgender Employees Count)
_AUTO_DERIVED_LABELS = {
    "agm day",
    "agm date in words",
    "authorized capital description",
    "issued capital description",
    "subscribed and paid-up capital",
    "transgender employees count",
}


def _scan_kv_required(ws, tab_name, key_col=2, val_col=3, formula_cells=None):
    """Walk a key/value tab. Every label ending in '*' must have a non-blank value.
    Cells that contain an Excel formula, OR whose label appears in
    _AUTO_DERIVED_LABELS (loader fills them in), are treated as filled."""
    formula_cells = formula_cells or set()
    missing = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=key_col).value
        if not label:
            continue
        label_str = str(label).strip()
        if not label_str.endswith("*"):
            continue
        if (ws.title, r, val_col) in formula_cells:
            continue
        if label_str.rstrip(" *").strip().lower() in _AUTO_DERIVED_LABELS:
            continue
        value = ws.cell(row=r, column=val_col).value
        if _is_blank(value):
            missing.append((tab_name, r, label_str.rstrip(" *")))
    return missing


def _has_any_data_row(ws, name_col, hdr_row=4, max_rows=50):
    for r in range(hdr_row + 1, hdr_row + 1 + max_rows):
        v = ws.cell(row=r, column=name_col).value
        if not _is_blank(v):
            return True
    return False


def _read_toggles(wb):
    """Read Toggles into a {key: 'YES'|'NO'} dict. Missing keys default to NO."""
    out = {}
    if "Toggles" not in wb.sheetnames:
        return out
    ws = wb["Toggles"]
    for r in range(1, ws.max_row + 1):
        k = ws.cell(row=r, column=2).value
        v = ws.cell(row=r, column=3).value
        if not k:
            continue
        k_str = str(k).strip()
        if k_str.startswith("▸") or k_str == "Toggle":
            continue
        out[k_str] = (str(v).strip().upper() if not _is_blank(v) else "NO")
    return out


def _is_yes(toggles, key):
    return toggles.get(key, "NO").upper() == "YES"


def find_missing(path):
    """Return [(tab, row_or_'—', label)] for every missing required field."""
    wb = load_workbook(path, data_only=True)
    # Also load WITHOUT data_only to detect formula-driven cells (cached value is
    # None when Excel hasn't recomputed yet; we don't want to flag those).
    wb_formulas = load_workbook(path, data_only=False)
    formula_cells = _collect_formula_cells(wb_formulas)
    missing = []

    if "Company" in wb.sheetnames:
        missing += _scan_kv_required(wb["Company"], "Company", formula_cells=formula_cells)

    toggles = _read_toggles(wb)

    # Auditor block is only mandatory when we're producing auditor content.
    auditor_needed = _is_yes(toggles, "AuditorReappointment") or _is_yes(toggles, "AuditorReportInDirReport")
    if auditor_needed and "Auditor" in wb.sheetnames:
        missing += _scan_kv_required(wb["Auditor"], "Auditor", formula_cells=formula_cells)

    if "Directors" in wb.sheetnames:
        if not _has_any_data_row(wb["Directors"], name_col=3):
            missing.append(("Directors", "—", "At least one director (Name in column C)"))

    if "Shareholders" in wb.sheetnames:
        if not _has_any_data_row(wb["Shareholders"], name_col=4):
            missing.append(("Shareholders", "—", "At least one shareholder (Name in column D)"))

    if "BoardMeetings" in wb.sheetnames:
        if not _has_any_data_row(wb["BoardMeetings"], name_col=2):
            missing.append(("BoardMeetings", "—", "At least one meeting date (column B)"))

    # Toggle-gated tabular tabs: only required when their toggle is YES.
    if _is_yes(toggles, "RPT_Required") and "RelatedParty" in wb.sheetnames:
        if not _has_any_data_row(wb["RelatedParty"], name_col=3):
            missing.append(("RelatedParty", "—",
                            "Toggles!RPT_Required=YES but no RelatedParty rows filled"))

    if _is_yes(toggles, "RegulariseAdditionalDirector"):
        # Director-tab YES flag OR DirChanges row is sufficient.
        regularise_via_directors = False
        if "Directors" in wb.sheetnames:
            ws = wb["Directors"]
            for r in range(5, 5 + 20):
                flag = ws.cell(row=r, column=7).value
                if flag and str(flag).strip().upper() in ("YES", "Y", "TRUE", "1"):
                    regularise_via_directors = True
                    break
        regularise_via_dirchanges = (
            "DirChanges" in wb.sheetnames
            and _has_any_data_row(wb["DirChanges"], name_col=2)
        )
        if not (regularise_via_directors or regularise_via_dirchanges):
            missing.append(("Directors", "—",
                            "Toggles!RegulariseAdditionalDirector=YES but no director marked Regularise=YES "
                            "(and DirChanges tab is empty)"))

    return missing


def _parse_excel_date(v):
    """Excel date-typed cells -> datetime.date; strings 'DD/MM/YYYY' -> date.
    Returns None on anything else."""
    from datetime import datetime, date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


_MONTHS = {m.lower(): i for i, m in enumerate(
    ["", "January", "February", "March", "April", "May", "June",
     "July", "August", "September", "October", "November", "December"])}


def _dates_mentioned_in_text(text):
    """Pull '27th April 2026' / '15 January 2025' / '27/04/2026' style dates
    out of free-text description columns. Returns list of date objects."""
    import re
    from datetime import date
    out = []
    if not text:
        return out
    s = str(text)
    # '27th April 2026' style
    pat1 = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+),?\s+(\d{4})",
        flags=re.IGNORECASE,
    )
    for m in pat1.finditer(s):
        day = int(m.group(1))
        month = _MONTHS.get(m.group(2).lower())
        year = int(m.group(3))
        if month and 1 <= day <= 31:
            try:
                out.append(date(year, month, day))
            except ValueError:
                pass
    # 'DD/MM/YYYY' style
    pat2 = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b")
    for m in pat2.finditer(s):
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            out.append(date(y, mo, d))
        except ValueError:
            pass
    return out


def find_warnings(path):
    """Return list of (severity, tab, row, label, message) for non-blocking
    issues — contradictions between toggles and data, date sanity, etc."""
    wb = load_workbook(path, data_only=True)
    warnings = []

    toggles = _read_toggles(wb)

    # Read Company-tab dates + employee counts for cross-checks
    company_dates = {}
    company_counts = {}
    if "Company" in wb.sheetnames:
        ws = wb["Company"]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=2).value
            if not label:
                continue
            norm = str(label).strip().rstrip("*").strip().lower()
            v = ws.cell(row=r, column=3).value
            # Match EXACT label first; partial matches drag in 'previous agm date'
            # / 'agm date in words' / 'agm day' / etc.
            if norm == "agm date":
                company_dates["agm_date"] = (_parse_excel_date(v), r)
            elif norm == "current fy end date":
                company_dates["fy_end"] = (_parse_excel_date(v), r)
            elif norm == "previous fy end date":
                company_dates["prev_fy_end"] = (_parse_excel_date(v), r)
            elif norm == "notice dispatch date":
                company_dates["notice_date"] = (_parse_excel_date(v), r)
            elif norm == "female employees count":
                try:
                    company_counts["female"] = (int(v), r) if v is not None else (0, r)
                except (TypeError, ValueError):
                    company_counts["female"] = (0, r)

    # 1. Maternity toggle vs Female employee count
    fem_count = company_counts.get("female", (0, None))[0]
    fem_row = company_counts.get("female", (0, None))[1]
    if fem_count > 0 and not _is_yes(toggles, "MaternityActApplicable"):
        warnings.append((
            "info", "Toggles", "—",
            "MaternityActApplicable",
            f"Female Employees Count = {fem_count} on Company tab (row {fem_row}). "
            f"Maternity Benefit Act is therefore applicable by law — the doc will "
            f"render the 'Company complies' branch regardless of this toggle.",
        ))

    # 2. ChangeOfNameDuringYear=YES but Old Name blank
    if _is_yes(toggles, "ChangeOfNameDuringYear") and "Company" in wb.sheetnames:
        ws = wb["Company"]
        old_name_blank = True
        old_name_row = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=2).value
            if label and "old name" in str(label).strip().lower():
                old_name_row = r
                v = ws.cell(row=r, column=3).value
                if v and str(v).strip():
                    old_name_blank = False
                break
        if old_name_blank:
            warnings.append((
                "warning", "Company", old_name_row or "—", "Old Name",
                "ChangeOfNameDuringYear is YES but Old Name is blank. "
                "The Change of Name paragraph (DR §28(1)) will have a placeholder.",
            ))

    # 3. HasSubsidiariesEtc=YES but no rows on a Subsidiaries tab (we never had one
    #    -> the doc renders an empty table). Surface that.
    if _is_yes(toggles, "HasSubsidiariesEtc"):
        warnings.append((
            "warning", "Toggles", "—", "HasSubsidiariesEtc",
            "Set to YES but there is no Subsidiaries tab in the workbook. "
            "DR §17 will render an empty placeholder table.",
        ))

    # 4. Director sanity checks (per-row contradictions)
    if "Directors" in wb.sheetnames:
        ws = wb["Directors"]
        fy_end = company_dates.get("fy_end", (None, None))[0]
        for r in range(5, min(ws.max_row, 30) + 1):
            name = ws.cell(row=r, column=3).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            appt = _parse_excel_date(ws.cell(row=r, column=5).value)
            cess = _parse_excel_date(ws.cell(row=r, column=6).value)
            regularise = str(ws.cell(row=r, column=7).value or "").strip().upper() in ("YES", "Y", "TRUE", "1")
            desc = ws.cell(row=r, column=8).value

            # 4a. Cessation < Appointment
            if appt and cess and cess < appt:
                warnings.append((
                    "warning", "Directors", r, name,
                    f"Date of Cessation ({cess.strftime('%d/%m/%Y')}) is BEFORE "
                    f"Date of Appointment ({appt.strftime('%d/%m/%Y')}).",
                ))

            # 4b. Regularise=YES but cessation is set
            if regularise and cess:
                warnings.append((
                    "warning", "Directors", r, name,
                    f"Marked Regularise=YES but has a Date of Cessation. "
                    f"You cannot regularise a director who has already exited.",
                ))

            # 4c. Appointment > FY end → excluded from "List of Directors as on FY end"
            if appt and fy_end and appt > fy_end:
                warnings.append((
                    "info", "Directors", r, name,
                    f"Appointment date {appt.strftime('%d/%m/%Y')} is AFTER current "
                    f"FY end {fy_end.strftime('%d/%m/%Y')}. This director is excluded "
                    f"from the 'List of Directors as on FY end' table and appears in "
                    f"the 'After end of financial year' sentence in DR §21 instead.",
                ))

            # 4d. Date mentioned in Description text doesn't match Appointment column.
            # Only flag when the description is talking about APPOINTMENT (looks for
            # "appointed", "appointment", or "board meeting" near a date). Exclude
            # any date that matches the cessation column — resignation descriptions
            # legitimately mention the cessation date.
            if appt and desc:
                desc_str = str(desc).lower()
                appointment_context = (
                    "appoint" in desc_str
                    or "board meeting" in desc_str
                    or "additional director" in desc_str
                )
                if appointment_context:
                    mentioned = [d for d in _dates_mentioned_in_text(desc) if d != cess]
                    if mentioned and appt not in mentioned:
                        flagged = mentioned[0]
                        warnings.append((
                            "warning", "Directors", r, name,
                            f"Date of Appointment column says "
                            f"{appt.strftime('%d/%m/%Y')} but the Description mentions "
                            f"{flagged.strftime('%d/%m/%Y')}. One of them is wrong.",
                        ))

    # 5. AGM date vs FY end (AGM should be after FY end and within 6 months)
    agm = company_dates.get("agm_date", (None, None))[0]
    fy_end = company_dates.get("fy_end", (None, None))[0]
    notice = company_dates.get("notice_date", (None, None))[0]
    if agm and fy_end and agm < fy_end:
        warnings.append((
            "warning", "Company",
            company_dates["agm_date"][1], "AGM Date",
            f"AGM Date ({agm.strftime('%d/%m/%Y')}) is BEFORE FY end "
            f"({fy_end.strftime('%d/%m/%Y')}). AGMs are held after the FY closes.",
        ))

    # 6. Notice dispatch < 21 days before AGM (Companies Act §101)
    if agm and notice:
        days = (agm - notice).days
        if days < 0:
            warnings.append((
                "warning", "Company",
                company_dates["notice_date"][1], "Notice Dispatch Date",
                f"Notice dispatched on {notice.strftime('%d/%m/%Y')} is AFTER the "
                f"AGM date ({agm.strftime('%d/%m/%Y')}). Notice must go out before "
                f"the AGM.",
            ))
        elif days < 21:
            warnings.append((
                "warning", "Company",
                company_dates["notice_date"][1], "Notice Dispatch Date",
                f"Notice dispatched only {days} days before the AGM. Section 101 of "
                f"the Companies Act requires a minimum 21-day notice.",
            ))

    return warnings


# Box-drawing constants kept ASCII-safe — Windows cmd.exe can render them.
_DIVIDER = "=" * 70
_SUBDIV  = "-" * 70


def validate_report(path):
    """Return a structured dict: {'errors': [...], 'warnings': [...], 'infos': [...]}.
    Each item is a dict with keys: tab, row, label, message."""
    errors = []
    for tab, row, label in find_missing(path):
        errors.append({"tab": tab, "row": row, "label": label,
                       "message": f"Required field is empty."})

    warnings = []
    infos = []
    for severity, tab, row, label, message in find_warnings(path):
        item = {"tab": tab, "row": row, "label": label, "message": message}
        (warnings if severity == "warning" else infos).append(item)

    return {"errors": errors, "warnings": warnings, "infos": infos}


def _print_group(title, marker, items):
    if not items:
        return
    print()
    print(f"  {marker} {title} ({len(items)})")
    print(f"  {_SUBDIV[2:]}")
    # Group by tab so the reader can fix one tab at a time
    by_tab = {}
    for it in items:
        by_tab.setdefault(it["tab"], []).append(it)
    for tab in sorted(by_tab):
        print(f"\n  [{tab}]")
        for it in by_tab[tab]:
            row = it["row"]
            label = it["label"]
            msg = it["message"]
            # Headline: row + label
            head = f"     Row {row:>3}  •  {label}" if row != "—" else f"     {label}"
            print(head)
            # Wrap the message at ~62 chars after a 7-space indent
            indent = "        "
            for line in _wrap(msg, width=62):
                print(indent + line)


def _wrap(text, width=62):
    """Tiny word-wrap that doesn't depend on textwrap module quirks."""
    words = str(text).split()
    line = ""
    out = []
    for w in words:
        if line and len(line) + 1 + len(w) > width:
            out.append(line)
            line = w
        else:
            line = (line + " " + w).strip()
    if line:
        out.append(line)
    return out


def print_report(report):
    """Pretty-print the structured report to the console."""
    n_err = len(report["errors"])
    n_warn = len(report["warnings"])
    n_info = len(report["infos"])

    print()
    print(_DIVIDER)
    print("  VALIDATION REPORT")
    print(_DIVIDER)

    if n_err == 0 and n_warn == 0 and n_info == 0:
        print("  ✓  All required fields are filled. No contradictions detected.")
        print(_DIVIDER)
        return

    _print_group("ERRORS — these MUST be fixed before generation",
                 "X", report["errors"])
    _print_group("WARNINGS — review before generating",
                 "!", report["warnings"])
    _print_group("INFO — auto-handled, just so you know",
                 "i", report["infos"])

    print()
    print(_SUBDIV)
    summary = f"  SUMMARY: {n_err} error{'s' if n_err != 1 else ''}"
    summary += f", {n_warn} warning{'s' if n_warn != 1 else ''}"
    summary += f", {n_info} info"
    print(summary)
    if n_err:
        print("  Fix the errors above in Master_Input.xlsx and run the app again.")
    else:
        print("  No blocking errors — generation can proceed.")
    print(_DIVIDER)


def validate(path):
    """Run validation, print a grouped report, and sys.exit(1) on errors.

    Backwards-compatible with the old call site: when there are no errors and
    no warnings, this is silent (just like before).
    """
    report = validate_report(path)
    n_err = len(report["errors"])
    n_warn = len(report["warnings"])
    n_info = len(report["infos"])

    # Silent on perfect runs to keep the CLI output clean — matches old behavior.
    if n_err == 0 and n_warn == 0 and n_info == 0:
        return

    print_report(report)
    if n_err:
        sys.exit(1)


if __name__ == "__main__":
    import os
    HERE = os.path.dirname(os.path.abspath(__file__))
    PROJECT_ROOT = os.path.dirname(HERE)
    report = validate_report(os.path.join(PROJECT_ROOT, "master", "Master_Input.xlsx"))
    print_report(report)
    if report["errors"]:
        sys.exit(1)
